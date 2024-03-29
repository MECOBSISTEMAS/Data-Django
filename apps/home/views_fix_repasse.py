# -*- encoding: utf-8 -*-
"""
Copyright (c) 2019 - present AppSeed.us
"""
#from typing import Self
from optparse import Values
from django.utils.text import slugify
from datetime import datetime, date, timedelta
from django import template
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, HttpRequest
from django.db.models import Case, Sum, When, F, Q, IntegerField, DecimalField, OuterRef, Subquery, Value, Max, Prefetch, DateField, QuerySet
from django.db.models.functions import Coalesce, Cast, TruncMonth
from django.template import loader
from django.urls import reverse
from django.shortcuts import render
from django.core.paginator import Paginator, Page, PageNotAnInteger, EmptyPage
from django.utils import timezone
from django.core import serializers
import random
import tempfile
import os
import json
import openpyxl
import asyncio
import calendar
import celery

from decimal import Decimal
from openpyxl.utils import get_column_letter
from requests import Request

from . import funcoes

from .existing_models import Contratos, ContratoParcelas, Pessoas, Eventos
#from .forms import CAD_ClienteForm, Calculo_RepasseForm
from .models import Calculo_Repasse, CadCliente, Debito, Credito, Taxa, RepasseRetido, Dado, RepasseAprovado, ParcelaTaxa


def filtrar_repasses(request:HttpRequest ,data_inicio:str, data_fim:str):
    data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
    data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
    context:dict = {}
    dados_dias = {}
    for i in range((data_fim_dt - data_inicio_dt).days + 1):
        data = data_inicio_dt + timedelta(days=i)
        dados_dias[f'{data.day}/{data.month}/{data.year}'] = None 
        """ Coalesce(
            Subquery(
                Dado.objects.filter(
                    id_vendedor=OuterRef('vendedor__id'),
                    dt_credito=data,
                    aprovada_para_repasse=False
                ).values('id_vendedor').annotate(
                    repasses_totais=Sum('repasses')
                ).values('repasses_totais'), output_field=DecimalField(max_digits=8, decimal_places=2)
            ),
        0,
        output_field=DecimalField(max_digits=8, decimal_places=2)) """
        
    context['repasses_clientes'] = CadCliente.objects.annotate(
        #**dados_dias,
        #some todos os repasses dentro da data de inicio e fim
        todos_os_repasses=Coalesce(
            Subquery(
                Dado.objects.filter(
                    id_vendedor=OuterRef('vendedor__id'),
                    dt_credito__range=(data_inicio, data_fim),
                    aprovada_para_repasse=False
                ).values('id_vendedor').annotate(
                    repasses_totais=Sum('repasses')
                ).values('repasses_totais'), 
                output_field=DecimalField(max_digits=8, decimal_places=2)
            ),
            0,
            output_field=DecimalField(max_digits=8, decimal_places=2)
            ) + 
            Coalesce(
                Subquery(
                    ContratoParcelas.objects.filter(
                        contratos__dt_contrato__range=(data_inicio, data_fim),
                        contratos__vendedores__id=OuterRef('vendedor__id'),
                        aprovada=False,
                        aprovada_para_repasse=False
                    ).values(
                        'contratos__vendedores__id',
                    ).annotate(
                        repasses_totais=Sum('vl_repasse')
                    ).values(
                        'repasses_totais'
                    ),
                    output_field=DecimalField(max_digits=8, decimal_places=2)
                ),
                0,
                output_field=DecimalField(max_digits=8, decimal_places=2)
            ),
        total_credito=Coalesce(
            Subquery(
                ParcelaTaxa.objects.filter(
                    id_vendedor=OuterRef('vendedor__id'),
                    data_aprovada__range=[data_inicio, data_fim],
                    aprovada=True,
                    aprovada_para_repasse=False,
                    deletada=False
                ).values('id_vendedor').annotate(total=Sum('repasse')).values('total'),
                output_field=DecimalField(max_digits=8, decimal_places=2),
            ),
            0,
            output_field=DecimalField(max_digits=8, decimal_places=2),
        )
        + Coalesce(
            Subquery(
                Credito.objects.filter(
                    cliente__id=OuterRef('vendedor__id'),
                    dt_creditado__range=[data_inicio, data_fim],
                    aprovada=True,
                    aprovada_para_repasse=False
                ).values('cliente__id').annotate(total=Sum('vl_credito')).values('total'),
                output_field=DecimalField(max_digits=8, decimal_places=2),
            ),
            0,
            output_field=DecimalField(max_digits=8, decimal_places=2),
        ),
        total_repasse_retido = Coalesce(
            Subquery(
                RepasseRetido.objects.filter(
                    cliente__id=OuterRef('vendedor__id'),
                    dt_rep_retido__range=(data_inicio, data_fim),
                    aprovada=True,
                    aprovada_para_repasse=False
                ).values('cliente__id').annotate(total=Sum('vlr_rep_retido')).values('total'),
                output_field=DecimalField(max_digits=8, decimal_places=2)
            ),
            0,
            output_field=DecimalField(max_digits=8, decimal_places=2)
        ),
        total_taxas = Coalesce(
            Subquery(
                Taxa.objects.filter(
                    cliente__id=OuterRef('vendedor__id'),
                    dt_taxa__range=(data_inicio, data_fim),
                    aprovada=True,
                    aprovada_para_repasse=False
                ).values('cliente__id').annotate(total=Sum('taxas')).values('total'),
                output_field=DecimalField(max_digits=8, decimal_places=2)
            ),
            0,
            output_field=DecimalField(max_digits=8, decimal_places=2)
        ),
        total_debitos = Coalesce(
            Subquery(
                Debito.objects.filter(
                    cliente__id=OuterRef('vendedor__id'),
                    dt_debitado__range=(data_inicio, data_fim),
                    aprovada=True,
                    aprovada_para_repasse=False
                ).values('cliente__id').annotate(total=Sum('vl_debito')).values('total'),
                output_field=DecimalField(max_digits=8, decimal_places=2),
            ),
            0,
            output_field=DecimalField(max_digits=8, decimal_places=2)
        ) + Coalesce(
            Subquery(
                ParcelaTaxa.objects.filter(
                    id_comprador=OuterRef('vendedor__id'),
                    data_aprovada__range=(data_inicio, data_fim),
                    aprovada=True,
                    aprovada_para_repasse=False,
                    deletada=False
                ).values('id').annotate(total=Sum('desconto_total')).values('total')[:1],
                output_field=DecimalField(max_digits=8, decimal_places=2),
        ),
        0,
        output_field=DecimalField(max_digits=8, decimal_places=2),
    ),
    total_repasses = F('total_credito') - F('total_taxas') - F('total_debitos') + F('total_repasse_retido') + F('todos_os_repasses')
    ).filter(Q(total_credito__gt=0) | Q(total_repasse_retido__gt=0) | Q(todos_os_repasses__gt=0)).values(
        'vendedor__id','vendedor__nome',
        'total_repasse_retido', 'total_credito',
        'total_taxas', 'total_debitos', 'total_repasses', 'todos_os_repasses',#*dados_dias.keys()
        ).order_by('vendedor__nome')
    context['dados_dias'] = dados_dias.keys()
    
    
    """ tbody = "<tr>"
    for repasse_cliente in context['repasses_clientes']:
        #tbody += f"<td><a class='btn btn-success' href='aprovar_repasse/{repasse_cliente['vendedor__id']}/{data_inicio}/{data_fim}/{repasse_cliente['total_repasse_retido']}/{repasse_cliente['total_credito']}/{repasse_cliente['total_taxas']}/{repasse_cliente['total_debitos']}/{repasse_cliente['total_repasses']}/' name='aprovar-repasse' id='aprovar-repasse'>Aprovar Repasses</a></td>"
        if request.user.is_superuser:
            tbody += f"<td><a name='aprovar-repasse' id='aprovar-repasse' class='btn btn-success' href='aprovar_repasse/{repasse_cliente['vendedor__id']}/{data_inicio}/{data_fim}'>Aprovar Repasses</a></td>"
        else:
            tbody += f"<td><a name='aprovar-repasse' id='aprovar-repasse' class='btn btn-success disabled' href='#sem-autorizacao'>Aprovar Repasses</a></td>"
            
        tbody += f"<td>{repasse_cliente['vendedor__id']}</td>"
        tbody += f"<td>{repasse_cliente['vendedor__nome']}</td>"
        tbody += f"<td>{repasse_cliente['total_repasse_retido']}</td>"
        tbody += f"<td>{repasse_cliente['todos_os_repasses']}</td>"
        tbody += f"<td>{repasse_cliente['total_credito']}</td>"
        tbody += f"<td>{repasse_cliente['total_taxas']}</td>"
        tbody += f"<td>{repasse_cliente['total_debitos']}</td>"
        tbody += f"<td>{repasse_cliente['total_repasses']}</td>"
        #tbody += f"<td>Total Repasses</td>"
        tbody += "</tr>"
    context['tbody'] = tbody """
    
    #faça o somatorio de creditos, debitos, repasses, taxas e repasses retidos e coloque no context como um subtotal
    context['total_credito'] = Decimal(sum([float(querie['total_credito']) for querie in context['repasses_clientes']])).quantize(Decimal('0.01'))
    context['total_debito'] = Decimal(sum([float(querie['total_debitos']) for querie in context['repasses_clientes']])).quantize(Decimal('0.01'))
    context['total_repasse_retido'] = Decimal(sum([float(querie['total_repasse_retido']) for querie in context['repasses_clientes']])).quantize(Decimal('0.01'))
    context['total_taxas'] = Decimal(sum([float(querie['total_taxas']) for querie in context['repasses_clientes']])).quantize(Decimal('0.01'))
    context['total_repasses'] = Decimal(sum([float(querie['total_repasses']) for querie in context['repasses_clientes']])).quantize(Decimal('0.01'))
    return context

class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        if isinstance(obj, date):
            return obj.strftime('%Y-%m-%d')
        if isinstance(obj, datetime):
            return obj.strftime('%Y-%m-%d %H:%M:%S')
        return super().default(obj)


def pegar_pessoa(pessoa_id):
    try:
        pessoa = Pessoas.objects.get(id=pessoa_id)
    except Exception as e:
        return HttpResponse('Pessoa não encontrada, erro: ' + str(e))
    return pessoa

@login_required(login_url="/login/")
def index(request):
    context = {'segment': 'index'}
    
    primeiro_dia_do_mes_atual = datetime.now().replace(day=1)
    decimo_quinto_dia_do_mes_atual = primeiro_dia_do_mes_atual + timedelta(days=14)
    ultimo_dia_do_mes_atual = primeiro_dia_do_mes_atual.replace(day=calendar.monthrange(primeiro_dia_do_mes_atual.year, primeiro_dia_do_mes_atual.month)[1])

    context['total_repasses_diario'] = Dado.objects.filter(
        dt_credito=date.today()
    ).aggregate(
        total_repasses=Sum('repasses')
    )['total_repasses'] or 0
    
    
    context['total_repasses_primeira_quinzena'] = Dado.objects.filter(
        dt_credito__range=[primeiro_dia_do_mes_atual - timedelta(days=1), decimo_quinto_dia_do_mes_atual]
    ).aggregate(
        total_repasses=Sum('repasses')
    )['total_repasses'] or 0
    
    
    context['total_repasses_segunda_quinzena'] = Dado.objects.filter(
        dt_credito__range=[decimo_quinto_dia_do_mes_atual, ultimo_dia_do_mes_atual]
    ).aggregate(
        total_repasses=Sum('repasses')
    )['total_repasses'] or 0

    # Encontrar o primeiro e o último dia da semana atual
    data_atual = date.today()
    primeiro_dia_da_semana = data_atual - timedelta(days=data_atual.weekday())
    ultimo_dia_da_semana = primeiro_dia_da_semana + timedelta(days=6)

    context['total_repasses_semanais'] = Dado.objects.filter(
        dt_credito__range=[primeiro_dia_da_semana, ultimo_dia_da_semana]
    ).aggregate(
        total_repasses=Sum('repasses')
    )['total_repasses'] or 0
    
    context['total_repasses_mensais'] = Dado.objects.filter(
        dt_credito__range=[primeiro_dia_do_mes_atual, ultimo_dia_do_mes_atual]
    ).aggregate(
        total_repasses=Sum('repasses')
    )['total_repasses'] or 0
    
    #?acima esta o codigo da construção das primeiras 5 cards
    
    # Obtenha o somatório dos "repasses" agrupados por mês do ano
    somatorio_por_mes_repasses_aprovados = RepasseAprovado.objects.annotate(
        mes=TruncMonth('data_aprovado', output_field=DateField())
    ).values('mes').annotate(
        total_repasses=Sum('repasses__repasses')
    )
    
    somatorio_por_mes_creditos = Credito.objects.filter(
        
        #filtrar pelo o ano atual
        
    ).annotate(
        mes=TruncMonth('dt_creditado', output_field=DateField())
    ).values('mes').annotate(
        total_credito=Sum('vl_credito')
    )
    
    somotario_por_mes_debitos = Debito.objects.filter(
        
        
    ).annotate(
        mes=TruncMonth('dt_debitado', output_field=DateField())
    ).values('mes').annotate(
        total_debito=Sum('vl_debito')
    )
    
    somotario_por_mes_repasses_retidos = RepasseRetido.objects.filter(
        
        
    ).annotate(
        mes=TruncMonth('dt_rep_retido', output_field=DateField())
    ).values('mes').annotate(
        total_repasse_retido=Sum('vlr_rep_retido')
    )
    
    somatorio_por_mes_taxas_totais = Taxa.objects.filter(
        
        
    ).annotate(
        mes=TruncMonth('dt_taxa', output_field=DateField())
    ).values('mes').annotate(
        total_taxas=Sum('taxas')
    )
    
    #?somatorio das taxas e seus tipos, encontrar uma forma melhor de fazer filtro e evitar repetição de codigo
    
    somatorio_por_mes_taxas_tbb = Taxa.objects.filter(
        tipo__icontains='TBB'
    ).annotate(
        mes=TruncMonth('dt_taxa', output_field=DateField())
    ).values('mes').annotate(
        total_taxas=Sum('taxas')
    )
    
    somatorio_por_mes_taxas_tec = Taxa.objects.filter(
        tipo__icontains='TEC'
    ).annotate(
        mes=TruncMonth('dt_taxa', output_field=DateField())
    ).values('mes').annotate(
        total_taxas=Sum('taxas')
    )
    
    somatorio_por_mes_taxas_tac = Taxa.objects.filter(
        tipo__icontains='TAC'
    ).annotate(
        mes=TruncMonth('dt_taxa', output_field=DateField())
    ).values('mes').annotate(
        total_taxas=Sum('taxas')
    )
    
    somatorio_por_mes_taxas_tcc = Taxa.objects.filter(
        tipo__icontains='TCC'
    ).annotate(
        mes=TruncMonth('dt_taxa', output_field=DateField())
    ).values('mes').annotate(
        total_taxas=Sum('taxas')
    )
    
    somatorio_por_mes_taxas_spc = Taxa.objects.filter(
        tipo__icontains='SPC'
    ).annotate(
        mes=TruncMonth('dt_taxa', output_field=DateField())
    ).values('mes').annotate(
        total_taxas=Sum('taxas')
    )
    
    somatorio_por_mes_taxas_confeccao_judicial = Taxa.objects.filter(
        tipo='Honorários Iniciais - Confecção de ação judicial'
    ).annotate(
        mes=TruncMonth('dt_taxa', output_field=DateField())
    ).values('mes').annotate(
        total_taxas=Sum('taxas')
    )
    
    somatorio_por_mes_taxas_honorarios_direto = Taxa.objects.filter(
        tipo='Honorários de recebimento direto'
    ).annotate(
        mes=TruncMonth('dt_taxa', output_field=DateField())
    ).values('mes').annotate(
        total_taxas=Sum('taxas')
    )
    
    
    somatorio_por_mes_taxas_honorarios_judiciais = Taxa.objects.filter(
        tipo='Honorários Judiciais'
    ).annotate(
        mes=TruncMonth('dt_taxa', output_field=DateField())
    ).values('mes').annotate(
        total_taxas=Sum('taxas')
    )
    
    
    def somatorio_por_mes(queryset:QuerySet, nome_do_campo_total:str, nome_da_consulta:str) -> dict:
        tabela_valores:dict = {
            nome_da_consulta: {}
        }

        for item in queryset:
        # Verifique se a chave 'mes' existe e tem valor antes de continuar
            if 'mes' in item and item['mes']:
                mes_do_ano = item['mes'].strftime('%B')  # Obtém o nome do mês (por extenso)
                total_repasses = item[nome_do_campo_total]
                mes_do_ano_int = item['mes'].month

                # Se ainda não existe a entrada para esse mês, crie-a
                tabela_valores[nome_da_consulta][mes_do_ano_int] = total_repasses
                """ if mes_do_ano not in tabela_valores[nome_da_consulta]:
                    tabela_valores[nome_da_consulta] = {mes_do_ano:total_repasses}
                else:
                    # Se já existe, some o valor ao total existente
                    tabela_valores[nome_da_consulta][mes_do_ano] = total_repasses """

        # Adicione outros meses com valor zero caso não existam dados para esses meses
        #for month in ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']:
        for month in range(1, 13):
            if month not in tabela_valores[nome_da_consulta]:
                tabela_valores[nome_da_consulta][month] = 0
        #organize as chaves do menor para o maior
        tabela_valores[nome_da_consulta] = dict(sorted(tabela_valores[nome_da_consulta].items()))
        return tabela_valores
    
    

    # Crie um dicionário vazio para tabelas_valores e inicialize a chave 'repasses_aprovados'
    tabelas_valores:dict = {}
    tabelas_valores.update(somatorio_por_mes(somatorio_por_mes_repasses_aprovados, 'total_repasses', 'Repasses Aprovados'))
    tabelas_valores.update(somatorio_por_mes(somatorio_por_mes_creditos, 'total_credito', 'Creditos'))
    tabelas_valores.update(somatorio_por_mes(somotario_por_mes_debitos, 'total_debito', 'Debitos'))
    tabelas_valores.update(somatorio_por_mes(somotario_por_mes_repasses_retidos, 'total_repasse_retido', 'Repasse Retido'))
    tabelas_valores.update(somatorio_por_mes(somatorio_por_mes_taxas_totais, 'total_taxas', 'Taxas Totais'))
    tabelas_valores.update(somatorio_por_mes(somatorio_por_mes_taxas_tbb, 'total_taxas', 'Taxas TBB'))
    tabelas_valores.update(somatorio_por_mes(somatorio_por_mes_taxas_tec, 'total_taxas', 'Taxas TEC'))
    tabelas_valores.update(somatorio_por_mes(somatorio_por_mes_taxas_tac, 'total_taxas', 'Taxas TAC'))
    tabelas_valores.update(somatorio_por_mes(somatorio_por_mes_taxas_tcc, 'total_taxas', 'Taxas TCC'))
    tabelas_valores.update(somatorio_por_mes(somatorio_por_mes_taxas_spc, 'total_taxas', 'Taxas SPC'))
    tabelas_valores.update(somatorio_por_mes(somatorio_por_mes_taxas_confeccao_judicial, 'total_taxas', 'Taxas Confeccao Judicial'))
    tabelas_valores.update(somatorio_por_mes(somatorio_por_mes_taxas_honorarios_direto, 'total_taxas', 'Taxas Honorarios Direto'))
    tabelas_valores.update(somatorio_por_mes(somatorio_por_mes_taxas_honorarios_judiciais, 'total_taxas', 'Taxas Honorarios Judiciais'))
    

    # Preencha o dicionário com os valores de repasses aprovados para cada mês do ano
    """ for item in somatorio_por_mes_repasses_aprovados:
        # Verifique se a chave 'mes' existe e tem valor antes de continuar
        if 'mes' in item and item['mes']:
            mes_do_ano = item['mes'].strftime('%B')  # Obtém o nome do mês (por extenso)
            total_repasses = item['total_repasses']

            # Se ainda não existe a entrada para esse mês, crie-a
            if mes_do_ano not in tabelas_valores['repasses_aprovados']:
                tabelas_valores['repasses_aprovados'][mes_do_ano] = total_repasses
            else:
                # Se já existe, some o valor ao total existente
                tabelas_valores['repasses_aprovados'][mes_do_ano] += total_repasses

    # Adicione outros meses com valor zero caso não existam dados para esses meses
    for month in ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']:
        if month not in tabelas_valores['repasses_aprovados']:
            tabelas_valores['repasses_aprovados'][month] = 0 """

    # Adicione o dicionário tabelas_valores ao contexto
    context['tabelas_valores'] = tabelas_valores
    
    if request.method == 'POST':
        if 'filtrar-valores' in request.POST:
            return HttpResponse('Filtrar valores')
    elif request.method == 'GET':
        if 'construir-dashs-ajax' in request.GET:
            return JsonResponse(
                data={
                    'message':'success',
                    'status': 200,
                    'table_values': json.dumps(tabelas_valores, cls=CustomJSONEncoder)
                }, safe=False, status=200)
        html_template = loader.get_template('home/index.html')
        return HttpResponse(html_template.render(context, request))


@login_required(login_url="/login/")
def pages(request):
    context:dict = {}

    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:
        request.session['serialized_data'] = None
        #request.session['prestacao_diaria_data'] = None
        load_template = request.path.split('/')[-1]

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        context['segment'] = load_template
        

        if load_template == 'cad_clientes_table_bootstrap.html':
            if request.method == 'POST':
                if 'novo-cliente-cadastro' in request.POST:
                    cliente_id = request.POST.get('cliente-id')
                    nome = request.POST.get('nome')
                    email = request.POST.get('email')
                    sim = request.POST.get('sim')
                    nao = request.POST.get('nao')
                    operacional = request.POST.get('operacional')
                    tcc = request.POST.get('tcc')
                    honorarios = request.POST.get('honorarios')
                    repasse_semanal = request.POST.get('valor_nao') == 'false'
                    
                    #!informar_repasse = request.POST.get('informar_repasse')
                    
                    if cliente_id == '' or cliente_id == None:
                        pessoa = Pessoas.objects.create(nome=nome, email=email)
                        CadCliente.objects.create(
                            vendedor = pessoa, sim=sim, nao=nao,
                            operacional=operacional, tcc=tcc,
                            honorarios=honorarios, repasse_semanal = not repasse_semanal
                        )
                    else:
                        try:
                            pessoa = Pessoas.objects.get(id=cliente_id)
                            return HttpResponse('Cliente já cadastrado')
                        except Pessoas.DoesNotExist:
                            #crie uma pessoa passando o nome e id_cliente e caso não exista o id passado pelo usuario crie um aleatorio
                            pessoa = Pessoas.objects.create(nome=nome, email=email, id=cliente_id)
                            CadCliente.objects.create(
                                vendedor = pessoa, sim=sim, nao=nao, 
                                operacional=operacional, tcc=tcc, 
                                honorarios=honorarios, repasse_semanal = not repasse_semanal
                            )
                        
            context['cad_clientes'] = CadCliente.objects.all()

        elif load_template == 'tbl_financeiro_nao_aprovados.html':
            if request.method == 'POST':
                data_inicio = request.POST.get('data-inicio')
                data_fim = request.POST.get('data-fim')
                context['data_inicio'] = data_inicio
                context['data_fim'] = data_fim
                financeiro_dias = {}
                for i in range((datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 1):
                    data = datetime.strptime(data_inicio, '%Y-%m-%d') + timedelta(days=i)
                    financeiro_dias[f'{data.day}/{data.month}/{data.year}'] = Sum(
                        Case(
                            When(dt_vencimento=data, then=F('repasse')),
                            default=0,
                            output_field=DecimalField(decimal_places=2, max_digits=14, validators=[]),
                        ),
                    )
                context['dias'] = list(range(1, (datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 2))
                context['financeiro_dias'] = financeiro_dias
                context['valores_financeiros'] = ParcelaTaxa.objects.filter(
                    data_aprovada__range=[data_inicio, data_fim],
                    aprovada=False,
                ).values('id_vendedor', 'vendedor').annotate(
                    **financeiro_dias,
                )
                context['soma_valores_financeiro'] = ParcelaTaxa.objects.filter(
                    data_aprovada__range=[data_inicio, data_fim], aprovada=False
                ).aggregate(
                    total_valor=Sum(Coalesce('valor', 0, output_field=DecimalField(decimal_places=2, max_digits=12))),
                    total_tcc=Sum(Coalesce('tcc', 0,output_field=DecimalField(decimal_places=2, max_digits=12))),
                    total_desconto=Sum(Coalesce('desconto_total', 0,output_field=DecimalField(decimal_places=2, max_digits=12))),
                    total_honorarios=Sum(Coalesce('honorarios', 0,output_field=DecimalField(decimal_places=2, max_digits=12))),
                    total_repasse=Sum(Coalesce('repasse', 0,output_field=DecimalField(decimal_places=2, max_digits=12))),
                )
                context['valores_financeiro_filtro'] = ParcelaTaxa.objects.filter(dt_vencimento__range=[data_inicio, data_fim])
                
                tbody = ""
                for valor_financeiro in context['valores_financeiros']:
                    tbody += "<tr>"
                    tbody += f"<td>{valor_financeiro['vendedor']}</td>"
                    #tbody += "<td>Teste</td>"
                    #!tbody += f"<td>{valor_financeiro['id_contrato']}</td>"
                    for dia in financeiro_dias:
                        valor_financeiro_dia = valor_financeiro[f'{dia}']
                        tbody += f"<td>{valor_financeiro_dia}</td>"
                    tbody += "</tr>"
                context['tbody_repasses'] = tbody
                
                context['tcc_filtro'] = ParcelaTaxa.objects.filter(
                    data_aprovada__range=[data_inicio, data_fim], tcc__isnull=False, aprovada=False,
                    ).values('id_vendedor', 'vendedor').annotate(
                        **funcoes.construir_dias_filtro(data_inicio, data_fim, 'tcc'),
                    )
                tbody = ""
                for tcc in context['tcc_filtro']:
                    tbody += "<tr>"
                    tbody += f"<td>{tcc['vendedor']}</td>"
                    for dia in financeiro_dias:
                        tcc_dia = tcc[f'{dia}']
                        tbody += f"<td>{tcc_dia}</td>"
                    tbody += "</tr>"
                context['tbody_tcc'] = tbody
                
                context['desconto_total_filtro'] = ParcelaTaxa.objects.filter(
                    data_aprovada__range=[data_inicio, data_fim], desconto_total__isnull=False, aprovada=False,
                    ).values('id_comprador', 'comprador').annotate(
                        **funcoes.construir_dias_filtro(data_inicio, data_fim, 'desconto_total'),
                    )
                tbody = ""
                for desconto_total in context['desconto_total_filtro']:
                    tbody += "<tr>"
                    tbody += f"<td>{desconto_total['comprador']}</td>"
                    for dia in financeiro_dias:
                        desconto_total_dia = desconto_total[f'{dia}']
                        tbody += f"<td>{desconto_total_dia}</td>"
                    tbody += "</tr>"
                context['tbody_desconto_total'] = tbody
                
                context['honorarios_filtro'] = ParcelaTaxa.objects.filter(
                    data_aprovada__range=[data_inicio, data_fim], honorarios__isnull=False, aprovada=False,
                    ).values('id_contrato', 'comprador', 'vendedor').annotate(
                        **funcoes.construir_dias_filtro(data_inicio, data_fim, 'honorarios'),
                    )
                tbody = ""
                for honorarios in context['honorarios_filtro']:
                    tbody += "<tr>"
                    tbody += f"<td>{honorarios['vendedor']}</td>"
                    for dia in financeiro_dias:
                        honorarios_dia = honorarios[f'{dia}']
                        tbody += f"<td>{honorarios_dia}</td>"
                    tbody += "</tr>"
                context['tbody_honorarios'] = tbody

        elif load_template == 'cob_diario_planilhas.html':
            if request.method == "POST":
                data_inicio = request.POST.get('data_inicio')  # 2022-08-01:str
                data_fim = request.POST.get('data_fim')  # 2022-08-21:str
                context['sql'] =  list(Dado.objects \
                    .filter(dt_credito__range=(data_inicio, data_fim)) \
                    .values('id_vendedor', 'id_contrato', 'vendedor', 'id_comprador', 'comprador',
                            'parcelas_contrato', 'vl_pago', 'dt_vencimento', 'dt_credito', 'banco', 
                            'contrato', 'evento', 'deposito', 'calculo', 'taxas', 'adi', 'me', 'op', 
                            'repasses', 'comissao') \
                    .annotate(total_repasse=Sum('repasses')) \
                    .values('id_vendedor', 'id_contrato', 'vendedor', 'id_comprador', 'comprador', 
                            'parcelas_contrato', 'vl_pago', 'dt_vencimento', 'dt_credito', 'banco', 
                            'contrato', 'evento', 'deposito', 'calculo', 'taxas', 'adi', 'me', 'op', 
                            'repasses', 'comissao', 'total_repasse', 'id'))
                request.session['serialized_data'] = json.dumps(context['sql'], cls=CustomJSONEncoder)
        
        elif load_template == 'tbl_prestacao_diaria.html':
            if request.method == 'POST':
                if 'exportar-xlsx' in request.POST:
                    if not request.session.get('prestacao_diaria_data'):
                        return HttpResponse('Nenhum dado encontrado para exportar')
                    return funcoes.exportar_planilha_prestacao_diaria(request)
                elif 'exportar-pdf' in request.POST:
                    return funcoes.gerar_arquivo_pdf(request, context)
                elif 'filtrar-prestacao-diaria' in request.POST:
                    #bancos:str = request.POST.get('bancos')
                    data = request.POST.get('data')
                    context['data_consultada'] = data
                    context['repasses_semanais'] = CadCliente.objects.filter(
                        repasse_semanal=True,
                    ).annotate(
                        total_repasses=Coalesce(
                            Subquery(
                                Dado.objects.filter(
                                    dt_credito=data,
                                    id_vendedor=OuterRef('vendedor_id')
                                ).values('id_vendedor').annotate(
                                    soma_repasses=Sum('repasses')
                                ).values('soma_repasses')
                            ),
                            0,
                            output_field=DecimalField()
                        )
                    ).values('total_repasses', 'vendedor__nome', 'vendedor__id')
                    
                    context['repasses'] = (
                        Dado.objects
                        .filter(dt_credito=data)
                        .values('id_contrato', 'vendedor')
                        .annotate(
                            repasses=Sum('repasses')
                        )
                        .order_by('vendedor')
                    )
                        
                    
                    context['valores_pagos_honorarios'] = Dado.objects.filter(dt_credito=data).aggregate(
                            valores_pagos=Sum('vl_pago'),
                            honorarios=Sum('me')
                        )
                        
                    context['comissionistas_do_mes'] = Dado.objects.filter(
                        dt_credito=data,comissao__isnull=False
                        ).values('comissao').annotate(comissoes=Sum('op'))
                    
                    context['repasses_geral'] = Dado.objects.filter(
                            dt_credito=data,
                        ).aggregate(
                            repasses=Sum('repasses')
                        )
                        
                    context['taxas'] = float(Dado.objects.filter(dt_credito=data).aggregate(taxas=Sum('taxas'))['taxas'] or 0)
                    repasses_geral = float(context['repasses_geral']['repasses'] or 0)
                    context['repasses_semanais_vendedores_totais'] = (sum([float(querie['total_repasses']) for querie in context['repasses_semanais']]))
                    
                    repasses_geral_descontado = repasses_geral - context['repasses_semanais_vendedores_totais']
                    
                    context['repasses_geral_descontado'] = float("{:.2f}".format(repasses_geral_descontado))
                    
                    request.session['prestacao_diaria_data'] = {
                        'repasses_semanais': json.dumps(list(context['repasses_semanais']), cls=CustomJSONEncoder),
                        'comissionistas_do_mes': json.dumps(list(context['comissionistas_do_mes']), cls=CustomJSONEncoder),
                        'valores_pagos_honorarios': json.dumps(list(context['valores_pagos_honorarios'].values()), cls=CustomJSONEncoder),
                        #'comissionistas_do_mes': json.dumps(list(context['comissionistas_do_mes'].values()), cls=CustomJSONEncoder),
                        'repasses_geral_descontado': json.dumps(context['repasses_geral_descontado'], cls=CustomJSONEncoder),
                        'taxas': json.dumps(context['taxas'], cls=CustomJSONEncoder),
                    }

            
        
        
        elif load_template == 'registrar_contratos.html':
            if request.method == 'POST':
                if request.headers.get('Content-Type') == 'application/json':
                    body:dict = json.loads(request.body)
                    if 'buscar-vendedor' in body:
                        pessoa_id:int = body.get('pessoa_id', None)
                        try:    
                            pessoa = Pessoas.objects.get(id=pessoa_id)
                            print(f"POST: {request.POST} \nBODY: {body} \n{pessoa_id}")
                            return JsonResponse(
                                data={
                                'pessoa': serializers.serialize('json', [pessoa, ]),
                                },
                                status=200,
                            )
                        except Exception as error:
                            return JsonResponse(
                                data={
                                    'message': 'error',
                                    'error': str(error),
                                },
                                status=404,
                            )
            pass

        elif load_template == 'tbl_bootstrap.html':
            if request.method == 'POST':
                if 'filtrar-primeira-quinzena' in request.POST:
                    # Filtrar primeira quinzena
                    today = datetime.now().date()
                    first_day_of_month = today.replace(day=1)
                    fifteenth_day_of_month = today.replace(day=15)

                    data_inicio = first_day_of_month.strftime('%Y-%m-%d')
                    data_fim = fifteenth_day_of_month.strftime('%Y-%m-%d')
                    context.update(filtrar_repasses(request, data_inicio, data_fim))
                    #context = filtrar_repasses(request, data_inicio, data_fim)

                elif 'filtrar-segunda-quinzena' in request.POST:
                    # Filtrar segunda quinzena
                    today = datetime.now().date()
                    sixteenth_day_of_month = today.replace(day=16)
                    last_day_of_month = today.replace(day=1) + timedelta(days=32)
                    last_day_of_month = last_day_of_month.replace(day=1) - timedelta(days=1)

                    data_inicio = sixteenth_day_of_month.strftime('%Y-%m-%d')
                    data_fim = last_day_of_month.strftime('%Y-%m-%d')
                    context.update(filtrar_repasses(request, data_inicio, data_fim))

                elif 'filtrar-tabela-repasse-cliente' in request.POST:
                    data_inicio = request.POST.get('data-inicio')
                    data_fim = request.POST.get('data-fim')
                    #context = filtrar_repasses(request, data_inicio, data_fim)
                    context.update(filtrar_repasses(request, data_inicio, data_fim))

                context['data_inicio'] = request.POST.get('data-inicio')
                context['data_fim'] = request.POST.get('data-fim')

            
                        
        elif load_template == 'tbl_repasses_aprovados.html':
            if request.method == 'POST':
                if "filtrar-repasses-aprovados" in request.POST:
                    context['data_inicio'] = request.POST.get('data-inicio')
                    context['data_fim'] = request.POST.get('data-fim')
                    context['repasses_aprovados'] = RepasseAprovado.objects.filter(
                        data_aprovado__range=[context['data_inicio'], context['data_fim']]
                    )
            #context['repasses_aprovados'] = RepasseAprovado.objects.all()
                


        elif load_template == 'tbl_resultado_financeiro.html':
            if request.method == 'POST':
                aprovacao = request.POST.get('aprovacao') == "aprovada"
                data_inicio = request.POST.get('data-inicio')
                data_fim = request.POST.get('data-fim')
                context['data_inicio'] = data_inicio
                context['data_fim'] = data_fim
                financeiro_dias = {}
                for i in range((datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 1):
                    data = datetime.strptime(data_inicio, '%Y-%m-%d') + timedelta(days=i)
                    financeiro_dias[f'{data.day}/{data.month}/{data.year}'] = Sum(
                        Case(
                            When(dt_vencimento=data, then=F('repasse')),
                            default=0,
                            output_field=DecimalField(decimal_places=2, max_digits=14, validators=[]),
                        ),
                    )
                    
                taxas_dias = {}
                for i in range((datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 1):
                    data = datetime.strptime(data_inicio, '%Y-%m-%d') + timedelta(days=i)
                    taxas_dias[f'{data.day}/{data.month}/{data.year}'] = Sum(
                        Case(
                            When(dt_taxa=data, then=F('taxas')),
                            default=0,
                            output_field=DecimalField(decimal_places=2, max_digits=14, validators=[]),
                        ),
                    )
                
                def construir_tbody_dinamico(dicionario:dict, nome_do_campo:str, dias:dict):
                    tbody = ""
                    for item in dicionario:
                        tbody += "<tr>"
                        tbody += f"<td>{item[nome_do_campo]}</td>"
                        for dia in dias:
                            tbody += f"<td>{item[f'{dia}']}</td>"
                        tbody += "</tr>"
                    return tbody
                
                
                context['financeiro_dias'] = financeiro_dias
                context['taxas_dias'] = taxas_dias
                context['valores_financeiros'] = ParcelaTaxa.objects.filter(
                    data_aprovada__range=[data_inicio, data_fim],
                    aprovada=aprovacao,
                ).values('id_vendedor', 'vendedor').annotate(
                    **financeiro_dias,
                )
                context['soma_valores_financeiro'] = ParcelaTaxa.objects.filter(
                    data_aprovada__range=[data_inicio, data_fim], aprovada=aprovacao
                ).aggregate(
                    total_valor=Sum(Coalesce('valor', 0, output_field=DecimalField(decimal_places=2, max_digits=12))),
                    total_tcc=Sum(Coalesce('tcc', 0,output_field=DecimalField(decimal_places=2, max_digits=12))),
                    total_desconto=Sum(Coalesce('desconto_total', 0,output_field=DecimalField(decimal_places=2, max_digits=12))),
                    total_honorarios=Sum(Coalesce('honorarios', 0,output_field=DecimalField(decimal_places=2, max_digits=12))),
                    total_repasse=Sum(Coalesce('repasse', 0,output_field=DecimalField(decimal_places=2, max_digits=12))),
                )
                context['valores_financeiro_filtro'] = ParcelaTaxa.objects.filter(dt_vencimento__range=[data_inicio, data_fim])
                
                        
                
                context['tbody_repasses_parcelas'] = construir_tbody_dinamico(context['valores_financeiros'], 'vendedor', financeiro_dias)
                
                context['tcc_filtro'] = ParcelaTaxa.objects.filter(
                    data_aprovada__range=[data_inicio, data_fim], tcc__isnull=False, aprovada=aprovacao,
                    ).values('id_vendedor', 'vendedor').annotate(
                        **funcoes.construir_dias_filtro(data_inicio, data_fim, 'tcc'),
                    )
                context['tbody_tcc_parcelas'] = construir_tbody_dinamico(context['tcc_filtro'], 'vendedor', financeiro_dias)
                
                context['desconto_total_filtro'] = ParcelaTaxa.objects.filter(
                    data_aprovada__range=[data_inicio, data_fim], desconto_total__isnull=False, aprovada=aprovacao,
                    ).values('id_comprador', 'comprador').annotate(
                        **funcoes.construir_dias_filtro(data_inicio, data_fim, 'desconto_total'),
                    )
                context['tbody_desconto_total_parcelas'] = construir_tbody_dinamico(context['desconto_total_filtro'], 'comprador', financeiro_dias)
                
                context['honorarios_filtro'] = ParcelaTaxa.objects.filter(
                    data_aprovada__range=[data_inicio, data_fim], honorarios__isnull=False, aprovada=aprovacao,
                    ).values('id_contrato', 'comprador', 'vendedor').annotate(
                        **funcoes.construir_dias_filtro(data_inicio, data_fim, 'honorarios'),
                    )
                context['tbody_honorarios_parcelas'] = construir_tbody_dinamico(context['honorarios_filtro'], 'vendedor', financeiro_dias)
                
                #filtro para as taxas
                
                context['taxa_tbb_filtro'] = Taxa.objects.filter(
                    tipo__icontains='TBB', dt_taxa__range=[data_inicio, data_fim], aprovada=aprovacao,
                ).values('cliente__nome').annotate(
                    **taxas_dias,
                )
                
                context['tbody_taxa_tbb_filtro'] = construir_tbody_dinamico(context['taxa_tbb_filtro'], 'cliente__nome', taxas_dias)
                
                context['taxa_tec_filtro'] = Taxa.objects.filter(
                    tipo__icontains='TEC', dt_taxa__range=[data_inicio, data_fim], aprovada=aprovacao,
                ).values('cliente__nome').annotate(
                    **taxas_dias,
                )
                
                context['tbody_taxa_tec_filtro'] = construir_tbody_dinamico(context['taxa_tec_filtro'], 'cliente__nome', taxas_dias)

                
                context['taxa_tac_filtro'] = Taxa.objects.filter(
                    tipo__icontains='TAC', dt_taxa__range=[data_inicio, data_fim], aprovada=aprovacao,
                ).values('cliente__nome').annotate(
                    **taxas_dias,
                )
                
                context['tbody_taxa_tac_filtro'] = construir_tbody_dinamico(context['taxa_tac_filtro'], 'cliente__nome', taxas_dias)
                
                context['taxa_tcc_filtro'] = Taxa.objects.filter(
                    tipo__icontains='TCC', dt_taxa__range=[data_inicio, data_fim], aprovada=aprovacao,
                ).values('cliente__nome').annotate(
                    **taxas_dias,
                )
                context['tbody_taxa_tcc_filtro'] = construir_tbody_dinamico(context['taxa_tcc_filtro'], 'cliente__nome', taxas_dias)
                
                
                context['taxa_spc_filtro'] = Taxa.objects.filter(
                    tipo__icontains='SPC', dt_taxa__range=[data_inicio, data_fim], aprovada=aprovacao,
                ).values('cliente__nome').annotate(
                    **taxas_dias,
                )
                context['tbody_taxa_spc_filtro'] = construir_tbody_dinamico(context['taxa_spc_filtro'], 'cliente__nome', taxas_dias)
                
                
                context['taxa_honorarios_iniciais_filtro'] = Taxa.objects.filter(
                    tipo='Honorários Iniciais - Confecção de ação judicial', dt_taxa__range=[data_inicio, data_fim], aprovada=aprovacao,
                ).values('cliente__nome').annotate(
                    **taxas_dias,
                )
                
                context['tbody_taxa_honorarios_iniciais_filtro'] = construir_tbody_dinamico(context['taxa_honorarios_iniciais_filtro'], 'cliente__nome', taxas_dias)
                
                context['taxa_honorarios_judiciais_filtro'] = Taxa.objects.filter(
                    tipo='Honorários Judiciais', dt_taxa__range=[data_inicio, data_fim], aprovada=aprovacao,
                ).values('cliente__nome').annotate(
                    **taxas_dias,
                )
                
                context['tbody_taxa_honorarios_judiciais_filtro'] = construir_tbody_dinamico(context['taxa_honorarios_judiciais_filtro'], 'cliente__nome', taxas_dias)
                
                context['taxa_honorarios_de_recebimento_direto_filtro'] = Taxa.objects.filter(
                    tipo='Honorários de recebimento direto', dt_taxa__range=[data_inicio, data_fim], aprovada=aprovacao,
                ).values('cliente__nome').annotate(
                    **taxas_dias,
                )
                
                context['tbody_taxa_honorarios_de_recebimento_direto_filtro'] = construir_tbody_dinamico(context['taxa_honorarios_de_recebimento_direto_filtro'], 'cliente__nome', taxas_dias)
                
                    
        elif load_template == 'tbl_taxas_aprovadas.html':
            if request.method == 'POST':
                if 'filtrar-parcela-taxa-aprovada' in request.POST:
                    data_inicio = request.POST.get('data-inicio')
                    data_fim = request.POST.get('data-fim')
                    context['parcelas_taxas_aprovadas'] = ParcelaTaxa.objects.filter(data_aprovada__range=(data_inicio, data_fim) , aprovada=True)
                    context['data_inicio'] = data_inicio
                    context['data_fim'] = data_fim
        
        elif load_template == 'pessoa_info.html':
            if request.method == 'POST':
                pass
            context['pessoa'] = Pessoas.objects.get(id=733 or None)
                
            

        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:
        pass

        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    """ except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request)) """





def upload_planilha_quinzenal(request, *args, **kwargs):
    if request.method == 'POST':
        planilha = request.FILES.get('docpicker')
        return HttpResponse(planilha)
    return HttpResponseRedirect('/tbl_credito_cessao.html')

def download_planilha_repasses(request, *args, **kwargs):
    return HttpResponse('configuração de download em andamento')
    if request.session.get('serialized_data') is None:
        return HttpResponse('Nenhum dado encontrado para download')
    with tempfile.TemporaryDirectory() as tmpdir:
        filepath = os.path.join(tmpdir, f'planilha_consulta_quinzenal_{slugify(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'ID Vendedor'
        sheet['B1'] = 'Nome Vendedor'
        sheet['C1'] = 'Valor Repasse Retido'
        
        # define a largura das primeiras colunas
        sheet.column_dimensions[get_column_letter(1)].width = 15
        sheet.column_dimensions[get_column_letter(2)].width = 30
        sheet.column_dimensions[get_column_letter(3)].width = 30
        
        for i, row in enumerate(json.loads(request.session.get('serialized_data'))):
            for j, value in enumerate(row):
                sheet.cell(row=i+2, column=j+1, value=row[value])
        workbook.save(filepath)
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(filepath)}"'
            return response

def download_planilha_cob(request, *args, **kwargs):
    if request.session.get('serialized_data') is None:
        return HttpResponse('Nenhum dado encontrado para download')
    with tempfile.TemporaryDirectory() as tmpdir:
        filepath = os.path.join(tmpdir, f'planilha_consulta_cob_{slugify(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        

        lista_header_consulta_sql = ['id_vendedor', 'id_contrato', 'vendedor', 'id_comprador', 'comprador',
                            'parcelas_contrato', 'vl_pago', 'dt_vencimento', 'dt_credito', 'banco',
                            'contrato', 'evento', 'deposito', 'calculo', 'taxas', 'adi', 'me', 'op',
                            'repasses', 'comissao', 'total_repasse', 'id']
        for i, value in enumerate(lista_header_consulta_sql):
            sheet[get_column_letter(i+1) + '1'] = str(value)
        
        for i, row in enumerate(json.loads(request.session.get('serialized_data'))):
            for j, header in enumerate(lista_header_consulta_sql):
                value = row.get(header, '')  # obtém o valor do dicionário usando a chave do cabeçalho
                sheet.cell(row=i+2, column=j+1, value=str(value))
        workbook.save(filepath)
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(filepath)}"'
            return response

def download_planilha_taxas_aprovadas(request, *args, **kwargs):
    taxas = Taxa.objects.filter(
        data_aprovada__range=[request.POST.get('data_inicio'), request.POST.get('data_fim')],
        aprovada=True,
    )
    with tempfile.TemporaryDirectory() as tmpdirname:
        filepath = os.path.join(tmpdirname, f'planilha_relatorio_parcelas_aprovadas_{slugify(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        cabecalho = ['Cliente', 'Data', 'Descrição', 'Valor', 'Tipo', 'Data de aprovação']
        
        for i, cab in enumerate(cabecalho):
            sheet.cell(row=1, column=i+1).value = cab
            
        row = 2  # Comece da segunda linha (após o cabeçalho)
        for taxa in taxas:
            sheet.cell(row=row, column=1, value= taxa.cliente.nome)
            sheet.cell(row=row, column=2, value = taxa.dt_taxa.strftime('%Y-%m-%d') if taxa.dt_taxa else '')
            sheet.cell(row=row, column=3, value = taxa.descricao if taxa.descricao else '')
            sheet.cell(row=row, column=4, value = str(taxa.taxas) if taxa.taxas else '')
            sheet.cell(row=row, column=5, value = taxa.tipo if taxa.tipo else '')
            sheet.cell(row=row, column=6, value = taxa.data_aprovada.strftime('%Y-%m-%d') if taxa.data_aprovada else '')
            row += 1

        workbook.save(filepath)
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(filepath)}"'
            return response
def download_planilha_taxas_aprovadas_quinzena(request, *args, **kwargs):
    """ <QueryDict: {
        'csrfmiddlewaretoken':
        ['Lgk7KvAiy1KBmZmHRPvTtXEX2I7zS3XBfy0Mxcft8elbdSbUbdrzJk2C3rzEdsab'], 
        'ano': ['2023'], 
        'mes': ['1'], #janeiro
        'quinzena': ['2'], 
        'exportar-planilha': ['']}> """
    #primeira quinzena: 1° ate o 14° dia do mes
    #segunda quinzena: 15° ate o ultimo dia do mes
    
    ano = int(request.POST.get('ano'))
    mes = int(request.POST.get('mes'))
    quinzena = int(request.POST.get('quinzena'))

    # Defina o primeiro dia do mês
    primeiro_dia_mes = date(ano, mes, 1)

    if quinzena == 1:
        # Primeira quinzena: 1° até o 14° dia do mês
        data_inicial = primeiro_dia_mes
        data_final = primeiro_dia_mes + timedelta(days=13)
    else:
        # Segunda quinzena: 15° até o último dia do mês
        ultimo_dia_mes = date(ano, mes % 12 + 1, 1) - timedelta(days=1)
        data_inicial = primeiro_dia_mes + timedelta(days=14)
        data_final = ultimo_dia_mes
    taxas = Taxa.objects.filter(
        data_aprovada__range=[data_inicial, data_final],
        aprovada=True,
    )
    with tempfile.TemporaryDirectory() as tmpdirname:
        filepath = os.path.join(tmpdirname, f'planilha_relatorio_parcelas_aprovadas_{slugify(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        cabecalho = ['Cliente', 'Data', 'Descrição', 'Valor', 'Tipo', 'Data de aprovação']
        
        for i, cab in enumerate(cabecalho):
            sheet.cell(row=1, column=i+1).value = cab
            
        row = 2  # Comece da segunda linha (após o cabeçalho)
        for taxa in taxas:
            sheet.cell(row=row, column=1, value= taxa.cliente.nome)
            sheet.cell(row=row, column=2, value = taxa.dt_taxa.strftime('%Y-%m-%d') if taxa.dt_taxa else '')
            sheet.cell(row=row, column=3, value = taxa.descricao if taxa.descricao else '')
            sheet.cell(row=row, column=4, value = str(taxa.taxas) if taxa.taxas else '')
            sheet.cell(row=row, column=5, value = taxa.tipo if taxa.tipo else '')
            sheet.cell(row=row, column=6, value = taxa.data_aprovada.strftime('%Y-%m-%d') if taxa.data_aprovada else '')
            row += 1

        workbook.save(filepath)
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(filepath)}"'
            return response

def upload_planilha_cob(request, *args, **kwargs):
    if request.method == 'POST':
        planilha = request.FILES.get('docpicker')
        #verificar se o arquivo é do tipo xlsx
        if planilha.name.endswith('.xlsx'):
            #arquivo esta recebendo com sucesso, azer os devidos tratamentos
            wb = openpyxl.load_workbook(planilha)
            #nesse arquivo de planilha, a primeira aba é a que contem os dados, e so existe uma aba
            sheet = wb.worksheets[0]
            linha = 0
            for row in sheet.iter_rows(values_only=True):
                if linha < 1:
                    linha += 1
                    continue
                linha += 1
                if (row[0] and row[1] and row[2]) == None:
                    break
            
            
        #arquivo esta recebendo com sucesso, azer os devidos tratamentos
        return HttpResponse(planilha)
    return HttpResponseRedirect('/form_elements.html')

def upload_planilha_cavalos_cob(request, *args, **kwargs):
    if request.method == 'POST':
        planilha = request.FILES.get('docpicker')
        if planilha is None:
            return HttpResponse('Nenhum arquivo selecionado')
        elif planilha.name.endswith('.xlsx'):
            pessoa_nula = Pessoas.objects.get(id=99999)
            erros:list[str] = []
            linhas_nulas = 0
            wb = openpyxl.load_workbook(planilha)
            cob = wb.active
            linha = 0
            for row in cob.iter_rows(values_only=True):
                if linha < 1:
                    linha += 1
                    continue
                if (row[0] and row[1] and row[2]) == (None or ''):
                    linhas_nulas += 1
                    if linhas_nulas == 2:
                        break
                    continue
                codigo_vendedor = row[0]
                contrato_id = row[1]
                nome_vendedor = row[2]
                nome_comprador = row[3]
                parcela_paga = row[4]
                contrato_parcelas = row[5]
                valor_parcela = row[6]
                dt_vencimento = row[7]#dia/mes/ando
                """ dt_vencimento_dt = datetime.combine(date(day=dt_vencimento.day, 
                    month=dt_vencimento.month, 
                    year=dt_vencimento.year), datetime.min.time()) """
                dt_credito = row[8]
                """ dt_credito_dt = datetime.combine(date(day=dt_credito.day,
                    month=dt_credito.month,
                    year=dt_credito.year), datetime.min.time()) """
                banco = row[9]
                produto_ou_contrato = row[10]
                evento = row[11]
                deposito = row[12]
                calc = row[13]
                taxas = row[14]
                adi = row[15]
                me = row[16]
                op = row[17]
                repasses = row[18]
                comissao = row[19]
                try:
                    vendedor = Pessoas.objects.get(id=codigo_vendedor)
                except Pessoas.DoesNotExist:
                    vendedor = Pessoas.objects.create(id=codigo_vendedor, nome=nome_vendedor)
                except Pessoas.MultipleObjectsReturned:
                    erros.append(f"O vendedor {nome_vendedor} possui mais de um cadastro. linha: {linha}")
                    continue
                try:
                    comprador = Pessoas.objects.get(nome=nome_comprador)
                except Pessoas.DoesNotExist:
                    comprador = Pessoas.objects.create(nome=nome_comprador, email=f"{nome_comprador}@gmail.com")
                except Pessoas.MultipleObjectsReturned or Exception as e:
                    erros.append(f"O comprador {nome_comprador} possui mais de um cadastro. linha: {linha}")
                    continue
                    
                if row[1] is not None and type(row[1]) == int:
                    if row[1] == 1:
                        eventos = Eventos.objects.create(nome=row[4], leiloeiro=pessoa_nula, dt_evento=date.today(), tipo="qualquer")
                        Contratos.objects.create(vendedor=vendedor, comprador=comprador, eventos=eventos, descricao=row[10], pessoas_id_inclusao=pessoa_nula)
                    else:
                        try:
                            contratos = Contratos.objects.get(id=row[1])
                            eventos = Eventos.objects.get(id=contratos.eventos.id)
                            eventos.nome = evento
                            contratos.eventos = eventos
                            contratos.descricao = produto_ou_contrato
                            eventos.save()
                            contratos.save()
                        except Contratos.DoesNotExist:
                            eventos = Eventos.objects.create(nome=evento, leiloeiro=Pessoas.objects.get(id=99999),
                                dt_evento=date.today(), tipo="qualquer")
                            contratos = Contratos.objects.create(id=contrato_id,
                                vendedor=vendedor, comprador=comprador,
                                descricao=row[10], eventos=eventos, pessoas_id_inclusao=pessoa_nula)
                else:
                    erros.append(f"O contrato {row[1]} não existe para o comprador: {row[3]} e vendedor: {row[2]}. linha: {linha}")
                    continue
                try:
                    contrato_parcelas = ContratoParcelas.objects.get(contratos=contratos, 
                        nu_parcela=int(str(row[5].split('/')[0][1:]))
                        )
                    contrato_parcelas.dt_vencimento = row[7]
                    contrato_parcelas.vl_parcela = row[6]
                    contrato_parcelas.dt_credito = row[8]
                    contrato_parcelas.save()
                except ContratoParcelas.DoesNotExist:
                    contrato_parcelas = ContratoParcelas.objects.create(
                        contratos=contratos,
                        dt_vencimento=row[7], vl_parcela=row[6], dt_credito=row[8],
                        nu_parcela=int(str(row[5].split('/')[0][1:]))
                    )
                except ContratoParcelas.MultipleObjectsReturned:
                    erros.append(f"O contrato {contratos.id} possui mais de uma parcela com o numero {row[5]}. linha: {linha}")
                    #!chances muito poucas de acontecer, mas se acontecer, o sistema vai pegar o primeiro objeto
                    pass
                try:
                    calculo_repasse = Calculo_Repasse.objects.get(contrato_parcelas=contrato_parcelas)
                    calculo_repasse.deposito = row[12]
                    calculo_repasse.calculo = row[13]
                    calculo_repasse.taxas = row[14]
                    calculo_repasse.adi = row[15]
                    calculo_repasse.me = row[16]
                    calculo_repasse.op = row[17]
                    calculo_repasse.repasses = row[18]
                    calculo_repasse.comissao = row[19]
                    calculo_repasse.dt_credito = contrato_parcelas.dt_credito
                    calculo_repasse.save()
                except Calculo_Repasse.DoesNotExist:
                    calculo_repasse = Calculo_Repasse.objects.create(contrato_parcelas=contrato_parcelas,
                        deposito=row[12], calculo=row[13], taxas=row[14], adi=row[15],
                        me=row[16], op=row[17], repasses=row[18], comissao=row[19], dt_credito=contrato_parcelas.dt_credito
                    )
                    
                linha += 1
            return HttpResponse('Planilha de cavalos recebida com sucesso, linhas lidas: {}, erros: {}'.format(linha,erros))
        else:
            return HttpResponse('Arquivo não é do tipo xlsx')

def upload_planilha_cad_clientes(request, *args, **kwargs):
    if request.method == 'POST':
        planilha = request.FILES.get('docpicker')
        if planilha is None:
            return HttpResponse('Nenhum arquivo selecionado')
        elif planilha.name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(planilha)
            cad_cliente = wb.active
            linha = 0
            erros:list[str] = []
            for row in cad_cliente.iter_rows(values_only=True):
                if linha < 1:
                    linha += 1
                    continue
                #* sistema de parada
                if (row[0]) == None:
                    break
                nome_vendedor = row[0]
                id_vendedor = row[1]
                sim = row[2]
                nao = row[3]
                operacional = row[4]
                tcc = row[5]
                honorarios = row[6]
                repasse_semanal = row[7]
                try:
                    vendedor = Pessoas.objects.get(id=id_vendedor)
                except Pessoas.DoesNotExist:
                    vendedor = Pessoas.objects.create(id=id_vendedor, nome=nome_vendedor)
                except Pessoas.MultipleObjectsReturned:
                    erros.append('Erro ao criar o vendedor, Multiplas Pessoas Encontradas: {}, linha: {}'.format(nome_vendedor, linha))
                    continue
                try:
                    cad_cliente = CadCliente.objects.get(vendedor__id=vendedor.id)
                    cad_cliente.sim = sim
                    cad_cliente.nao = nao
                    cad_cliente.operacional = operacional
                    cad_cliente.tcc = tcc
                    cad_cliente.honorarios = honorarios
                    cad_cliente.repasse_semanal = True if str(repasse_semanal) == 'SIM' else False
                    cad_cliente.save()
                except CadCliente.DoesNotExist:
                    cad_cliente = CadCliente.objects \
                        .create(vendedor=vendedor, 
                            sim=sim, nao=nao, operacional=operacional, 
                            tcc=tcc, honorarios=honorarios, 
                            repasse_semanal=True if repasse_semanal == 'SIM' else False)
                except CadCliente.MultipleObjectsReturned:
                    erros.append('Erro ao criar o cadastro do cliente, Multiplos CadCliente,s Encontrados: {}, linha: {}'.format(nome_vendedor, linha))
                    continue
                
            return HttpResponse('Planilha {} recebida com sucesso'.format(planilha.name))
        else:
            return HttpResponse('Arquivo não é do tipo xlsx')

def upload_planilha_parcelas_taxas(request, *args, **kwargs):
    if request.method == "POST":
        planilha = request.FILES.get('docpicker')
        if planilha is None:
            return HttpResponse('Nenhum arquivo selecionado')
        elif not planilha.name.endswith('.xlsx'):
            return HttpResponse('Arquivo não é do tipo xlsx')
        linhas_nulas = 0
        linhas = 0
        parcelas_criadas:int = 0
        erros:list[str] = []
        wb = openpyxl.load_workbook(planilha)
        segunda_quinzena = wb.active
        for row in segunda_quinzena.iter_rows(values_only=True):
            if linhas < 1:
                linhas += 1
                continue
            if row[0] == None or row[0] == '':
                break
            id_contrato = row[0]
            id_comprador = row[1]
            nome_comprador = row[2]
            id_vendedor = row[3]
            nome_vendedor = row[4]
            parcela = row[5]
            dt_vencimento = row[6]#datetime.strptime(row[6], "<%d/%m/%Y>").date()
            valor = row[7]
            tcc = row[8]
            ted = row[9]
            desconto_total = row[10]
            honorarios = row[11]
            repasse = row[12]
            try:
                parcela_taxa = ParcelaTaxa.objects.get(
                        id_contrato=id_contrato, id_comprador=id_comprador, 
                        id_vendedor=id_vendedor, parcela=parcela,dt_vencimento=dt_vencimento,
                    )
                parcela_taxa.valor = valor
                parcela_taxa.tcc = tcc
                #parcela_taxa.ted = ted (não há o campo ted no modelo ParcelaTaxa ainda)
                parcela_taxa.desconto_total = desconto_total
                parcela_taxa.honorarios = honorarios
                parcela_taxa.repasse = repasse
                parcela_taxa.save()
            except ParcelaTaxa.DoesNotExist:
                ParcelaTaxa.objects.create(
                    id_contrato=id_contrato, id_comprador=id_comprador, comprador=nome_comprador,
                    id_vendedor=id_vendedor, vendedor=nome_vendedor, parcela=parcela, dt_vencimento=dt_vencimento,
                    valor=valor, tcc=tcc, desconto_total=desconto_total, honorarios=honorarios, repasse=repasse
                )
                parcelas_criadas += 1
            except ParcelaTaxa.MultipleObjectsReturned:
                erros.append(f"Multiplos objetos retornados na linha: {linhas}, quantidade retornanda: {ParcelaTaxa.objects.filter(id_contrato=id_contrato, id_comprador=id_comprador, id_vendedor=id_vendedor, parcela=parcela,dt_vencimento=dt_vencimento).count()}")
            except Exception as e:
                erros.append(f"Erro na linha {linhas}, Exception Error:{e}")
            linhas += 1
        
        return HttpResponse("Planilha Recebida com sucesso, <br> linhas lidas: {} , <br> criadas: {}, <br> erros: {}".format(linhas, parcelas_criadas, erros))

def upload_planilha_taxas(request):
    if request.method == "POST":
        if not request.FILES.get('docpicker').name.endswith('.xlsx'):
            return HttpResponse("Arquivo não é do tipo .xlsx")
        linhas:int = 0
        linhas_nulas:int = 0
        taxas_criadas:int = 0
        erros:list[str] = []
        wb = openpyxl.load_workbook(request.FILES.get('docpicker'))
        planilha = wb.active
        for row in planilha.iter_rows(values_only=True):
            if linhas < 1:
                linhas += 1
                continue
            if row[0] == None or row[0] == '':
                linhas_nulas += 1
                if linhas_nulas > 1:
                    break
                continue
            linhas += 1
            id_vendedor = row[0]
            nome_vendedor = row[1]
            data_lancamento = row[2]
            tipo = row[3]
            valor = row[4]
            try:
                taxa = Taxa.objects.create(
                    cliente=Pessoas.objects.get(id=id_vendedor),
                    taxas=valor,
                    dt_taxa=data_lancamento,
                    tipo=tipo,
                )
                taxas_criadas += 1
            except Pessoas.DoesNotExist:
                erros.append(f"O vendedor {nome_vendedor} não existe, linha: {linhas}")
                continue
            except Exception as error:
                erros.append(f"Erro na linha {linhas}, Exception Error:{error}")
                continue
        mensagem = f"""
        Planilha Recebida com sucesso, <br>
        Taxas criadas: {taxas_criadas}, <br>
        erros: {erros}
        """
        return HttpResponse(mensagem)
    return HttpResponse("TIPO GET ?")


""" def upload_planilha_dados_brutos(request):
    if request.method == "POST":
        if request.FILES.get('docpicker') is None:
            return HttpResponse("Nenhum arquivo selecionado")
        if not request.FILES.get('docpicker').name.endswith('.xlsx'):
            return HttpResponse("Arquivo não é do tipo xlsx")
        planilha = request.FILES.get('docpicker')
        wb = openpyxl.load_workbook(planilha)
        cob = wb.active
        dados_criados = 0
        dados_modificados = 0
        linhas = 0
        erros = []

        existing_data = {}
        new_data = []
        update_data = []

        from datetime import datetime

        for row in cob.iter_rows(values_only=True):
            if linhas == 0:
                linhas += 1
                continue
            if not any(row):
                continue

            vendedor_id, contrato_id, nome_vendedor, comprador, parcela_paga, parcelas_contrato, valor, data_vencimento, data_credito, banco, contrato, evento, calc, taxas, adi, me, op, repasses, comissao = row

            if (
                not vendedor_id
                and not contrato_id
                and not parcela_paga
                and not comprador
                and not valor
                and not contrato
                and not evento
            ):
                continue

            if not all([vendedor_id, contrato_id, parcela_paga, comprador, valor, contrato, evento]):
                continue

            try:
                # Verifique se data_credito é uma data válida no formato YYYY-MM-DD
                data_credito = datetime.strptime(data_credito, "%Y-%m-%d")
            except ValueError:
                erros.append(f"Formato de data de crédito inválido na linha {linhas}: {data_credito}. Deve ser no formato YYYY-MM-DD.")
                continue

            key = (
                vendedor_id,
                contrato_id,
                parcela_paga,
                comprador,
                valor,
                contrato,
                evento,
                data_vencimento,
                data_credito,
            )

            if key in existing_data:
                dado = existing_data[key]
                dado.parcelas_contrato = parcelas_contrato
                dado.banco = banco
                dado.calculo = calc
                dado.taxas = taxas
                dado.adi = adi
                dado.me = me
                dado.op = op
                dado.repasses = repasses
                dado.comissao = comissao
                dado.vendedor = nome_vendedor
                dados_modificados += 1
            else:
                try:
                    dado = Dado.objects.get(
                        id_vendedor=vendedor_id,
                        id_contrato=contrato_id,
                        nu_parcela=parcela_paga,
                        comprador=comprador,
                        vl_pago=valor,
                        contrato=contrato,
                        evento=evento,
                        dt_vencimento=data_vencimento,
                        dt_credito=data_credito,
                    )
                    dado.parcelas_contrato = parcelas_contrato
                    dado.banco = banco
                    dado.calculo = calc
                    dado.taxas = taxas
                    dado.adi = adi
                    dado.me = me
                    dado.op = op
                    dado.repasses = repasses
                    dado.comissao = comissao
                    dado.vendedor = nome_vendedor
                    dados_modificados += 1
                except Dado.DoesNotExist:
                    new_data.append(
                        Dado(
                            id_vendedor=vendedor_id,
                            id_contrato=contrato_id,
                            vendedor=nome_vendedor,
                            comprador=comprador,
                            nu_parcela=parcela_paga,
                            parcelas_contrato=parcelas_contrato,
                            vl_pago=valor,
                            dt_vencimento=data_vencimento,
                            dt_credito=data_credito,
                            banco=banco,
                            contrato=contrato,
                            evento=evento,
                            calculo=calc,
                            taxas=taxas,
                            adi=adi,
                            me=me,
                            op=op,
                            repasses=repasses,
                            comissao=comissao,
                        )
                    )
                    dados_criados += 1

            existing_data[key] = dado

        if new_data:
            Dado.objects.bulk_create(new_data)
        if update_data:
            Dado.objects.bulk_update(update_data, ["parcelas_contrato", "banco", "calculo", "taxas", "adi", "me", "op", "repasses", "comissao"])

        erros_html = "<ul>"
        for erro in erros:
            erros_html += f"<li>{erro}</li>"
        erros_html += "</ul>"

        return HttpResponse(
            "Planilha recebida com sucesso  <br> Dados Criados: {} <br> Dados Modificados: {} <br> Erros: {}".format(
                dados_criados, dados_modificados, erros_html
            )
        )

    return HttpResponse("HTTP REQUEST") """



def upload_planilha_dados_brutos(request):
    if request.method == "POST":
        if request.FILES.get('docpicker') is None:
            return HttpResponse("Nenhum arquivo selecionado")
        if not request.FILES.get('docpicker').name.endswith('.xlsx'):
            return HttpResponse("Arquivo não é do tipo xlsx")
        planilha = request.FILES.get('docpicker')
        wb = openpyxl.load_workbook(planilha)
        cob = wb.active
        linhas_nulas = 0
        dados_modificados = 0
        dados_criados = 0
        linhas = 0
        erros:list[str] = []
        lista_de_dados_que_serao_criados:list = []
        for row in cob.iter_rows(values_only=True):
            if linhas < 1:
                linhas += 1
                continue
            if (row[0] == None or row[0] == '') and (row[1] == None or row[1] == ''):
                linhas_nulas += 1
                if linhas_nulas > 1:
                    break
                continue
            linhas += 1
            vendedor_id = row[0]
            contrato_id = row[1]
            nome_vendedor = row[2]
            comprador = row[3]
            parcela_paga = row[4]
            parcelas_contrato = row[5]
            valor = row[6]
            data_vencimento = row[7]
            data_credito = row[8]
            banco = row[9]
            contrato = row[10]
            evento = row[11]
            #deposito = row[12]
            calc = row[12]
            taxas = row[13]
            adi = row[14]
            me = row[15]
            op = row[16]
            repasses = row[17]
            comissao = row[18]
            
            try:
                dado = Dado.objects.get(
                    id_vendedor=vendedor_id, 
                    id_contrato=contrato_id, 
                    nu_parcela=parcela_paga, 
                    comprador=comprador, 
                    vl_pago=valor, 
                    contrato=contrato, 
                    evento=evento,
                    dt_vencimento=data_vencimento,
                    dt_credito=data_credito,
                )
                #dado.nu_parcela = parcela_paga
                dado.parcelas_contrato = parcelas_contrato
                #dado.dt_vencimento = data_vencimento
                #dado.dt_credito = data_credito
                dado.banco = banco
                #dado.contrato = contrato
                #dado.evento = evento
                #dado.deposito = deposito
                dado.calculo = calc
                dado.taxas = taxas
                dado.adi = adi
                dado.me = me
                dado.op = op
                dado.repasses = repasses
                dado.comissao = comissao
                dado.vendedor = nome_vendedor
                dado.save()
                dados_modificados += 1
            except Dado.DoesNotExist:
                lista_de_dados_que_serao_criados.append(Dado(
                id_vendedor=vendedor_id,
                id_contrato=contrato_id,
                vendedor=nome_vendedor,
                comprador=comprador,
                nu_parcela=parcela_paga,
                parcelas_contrato=parcelas_contrato,#(1/2) :str, pegar somente o primeiro digte antes da barra
                vl_pago=valor,
                dt_vencimento=data_vencimento,
                dt_credito=data_credito,
                banco=banco,
                contrato=contrato,
                evento=evento,
                #deposito=deposito,
                calculo=calc,
                taxas=taxas,
                adi=adi,
                me=me,
                op=op,
                repasses=repasses,
                comissao=comissao
            ))
                dados_criados += 1
            except Dado.MultipleObjectsReturned:
                erros.append(f"Foi encontrado mais de um dado com os mesmos dados na linha {linhas}\n chaves primarias de busca são: id_vendedor={vendedor_id}, id_contrato={contrato_id}, comprador={comprador}, vl_pago={valor}")
            except Exception as e:
                erros.append(f"Erro na linha {linhas}, Exception Error:{e}")
            
        Dado.objects.bulk_create(lista_de_dados_que_serao_criados)
        erros_html = "<ul>"
        for erro in erros:
            erros_html += f"<li>{erro}</li>"
        erros_html = "</ul>"
        return HttpResponse("Planilha recebida com sucesso <br> linhas lidas: {} <br> Dados Criados: {} <br> Dados Modificados : {} <br> erros: {}".format(linhas,dados_criados,dados_modificados, erros_html))
    return HttpResponse("HTTP REQUEST")

def aprovar_repasse(request, *args, **kwargs):
    data_inicio = kwargs.get('data_inicio')
    data_fim = kwargs.get('data_fim')
    pessoa = Pessoas.objects.get(id=kwargs.get('id_pessoa'))
    creditos = Credito.objects.filter(cliente__id=pessoa.id, dt_creditado__range=[data_inicio, data_fim], aprovada=True)
    debitos = Debito.objects.filter(cliente__id=pessoa.id, dt_debitado__range=[data_inicio, data_fim], aprovada=True)
    taxas = Taxa.objects.filter(cliente__id=pessoa.id, dt_taxa__range=[data_inicio, data_fim], aprovada=True)
    repasses = Dado.objects.filter(id_vendedor=pessoa.id, dt_credito__range=[data_inicio, data_fim])
    repasses_retidos = RepasseRetido.objects.filter(cliente__id=pessoa.id, dt_rep_retido__range=[data_inicio, data_fim], aprovada=True)
    parcelas_taxas = ParcelaTaxa.objects.filter(Q(id_vendedor=pessoa.id) | Q(id_comprador=pessoa.id), dt_vencimento__range=[data_inicio, data_fim], aprovada=True)
    repasse_aprovado = RepasseAprovado.objects.create(
        cliente=pessoa,
        data_inicial=data_inicio,
        data_final=data_fim
    )
    creditos.update(aprovada_para_repasse=True)
    debitos.update(aprovada_para_repasse=True)
    taxas.update(aprovada_para_repasse=True)
    repasses.update(aprovada_para_repasse=True)
    repasses_retidos.update(aprovada_para_repasse=True)
    parcelas_taxas.update(aprovada_para_repasse=True)
    
    repasse_aprovado.creditos.add(*creditos)
    repasse_aprovado.debitos.add(*debitos)
    repasse_aprovado.taxas.add(*taxas)
    repasse_aprovado.repasses_retidos.add(*repasses_retidos)
    repasse_aprovado.repasses.add(*repasses)
    repasse_aprovado.parcelas_taxas.add(*parcelas_taxas)
        
    repasse_aprovado.save()
    
    return JsonResponse({'message': 'Repasse aprovado com sucesso'}, status=200)
    return HttpResponseRedirect('/tbl_bootstrap.html')
    
    return JsonResponse(
        {
        'dados':list(dados[0]),
        'dados_dias':list(dados[1]),
        'data_inicio':data_inicial,
        'data_fim':data_final, 
        'status':200
        }
    )
    return JsonResponse({'dados': dados_nao_aprovados_json, 'data_inicial': data_inicial, 'data_final': data_final, 'status': 200})

def desaprovar_repasse(request, *args, **kwargs):
    repasse_aprovado = RepasseAprovado.objects.get(id=kwargs.get('repasse_aprovado_id'))
    repasse_aprovado.creditos.all().update(aprovada_para_repasse=False)
    repasse_aprovado.debitos.all().update(aprovada_para_repasse=False)
    repasse_aprovado.taxas.all().update(aprovada_para_repasse=False)
    repasse_aprovado.repasses_retidos.all().update(aprovada_para_repasse=False)
    repasse_aprovado.repasses.all().update(aprovada_para_repasse=False)
    repasse_aprovado.parcelas_taxas.all().update(aprovada_para_repasse=False)
    
    repasse_aprovado.delete()
    return HttpResponseRedirect('/tbl_repasses_aprovados.html')

def aprovar_parcela_taxa(request, *args, **kwargs):
    parcela_taxa = ParcelaTaxa.objects.get(id=kwargs.get('parcela_taxa_id'))
    data_inicio = kwargs.get('data_inicio')
    data_fim = kwargs.get('data_fim')
    parcela_taxa.aprovada = True
    #coloque a data de hoje no campo data_aprovada
    parcela_taxa.data_aprovada = datetime.now()
    parcela_taxa.save()
    
    return JsonResponse(
        {
            'parcelas_taxas': list(ParcelaTaxa.objects.filter(dt_vencimento__range=(data_inicio, data_fim), aprovada=False).values()),
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'status': 200,
        }
    )
    return HttpResponseRedirect('/tbl_parcela_taxas.html')

def desaprovar_parcela_taxa(request, *args, **kwargs):
    parcela_taxa_aprovada = ParcelaTaxa.objects.get(id=kwargs.get('parcela_taxa_id'))
    parcela_taxa_aprovada.aprovada = False
    parcela_taxa_aprovada.save()
    return HttpResponseRedirect('tbl_taxas_aprovadas.html')

def aprovar_taxa_manual(request, *args, **kwargs):
    data_inicio = kwargs.get('data_inicio')
    data_fim = kwargs.get('data_fim')
    print(data_inicio, data_fim)
    taxa = Taxa.objects.get(id=kwargs.get('taxa_id'))
    taxa.aprovada = True
    taxa.save()
    taxas_nao_aprovadas = Taxa.objects.filter(aprovada=False, dt_taxa__range=(data_inicio, data_fim))
    taxas_aprovadas = Taxa.objects.filter(aprovada=True, dt_taxa__range=(data_inicio, data_fim))
    return HttpResponseRedirect('/tbl_taxas.html')  
    return JsonResponse(
        status = 200,
        data = {
            'taxas_nao_aprovadas': list(taxas_nao_aprovadas.values()),
            'taxas_aprovadas': list(taxas_aprovadas.values()),
        },
    )

def desaprovar_taxa_manual(request, *args, **kwargs):
    data_inicio = kwargs.get('data_inicio')
    data_fim = kwargs.get('data_fim')
    taxa = Taxa.objects.get(id=kwargs.get('taxa_id'))
    taxa.aprovada = False
    taxa.save()
    taxas_nao_aprovadas = Taxa.objects.filter(aprovada=False, dt_taxa__range=(data_inicio, data_fim))
    taxas_aprovadas = Taxa.objects.filter(aprovada=True, dt_taxa__range=(data_inicio, data_fim))
    return HttpResponseRedirect('/tbl_taxas.html')
    return JsonResponse(
        status = 200,
        data = {
            'taxas_nao_aprovadas': list(taxas_nao_aprovadas.values()),
            'taxas_aprovadas': list(taxas_aprovadas.values()),
        },
    )

def aprovar_credito(request, *args, **kwargs):
    credito = Credito.objects.get(id=kwargs.get('credito_id'))
    credito.aprovada = True
    credito.save()
    return HttpResponseRedirect('/tbl_credito.html')

def desaprovar_credito(request, *args, **kwargs):
    credito = Credito.objects.get(id=kwargs.get('credito_id'))
    credito.aprovada = False
    credito.save()
    return HttpResponseRedirect('/tbl_credito.html')

def aprovar_debito(request, *args, **kwargs):
    debito = Debito.objects.get(id=kwargs.get('debito_id'))
    debito.aprovada = True
    debito.save()
    return HttpResponseRedirect('/tbl_debito.html')

def desaprovar_debito(request, *args, **kwargs):
    debito = Debito.objects.get(id=kwargs.get('debito_id'))
    debito.aprovada = False
    debito.save()
    return HttpResponseRedirect('/tbl_debito.html')

def aprovar_repasse_retido(request, *args, **kwargs):
    repasse_retido = RepasseRetido.objects.get(id=kwargs.get('repasse_retido_id'))
    repasse_retido.aprovada = True
    repasse_retido.save()
    return HttpResponseRedirect('/tbl_repasse_retido.html')

def desaprovar_repasse_retido(request, *args, **kwargs):
    repasse_retido = RepasseRetido.objects.get(id=kwargs.get('repasse_retido_id'))
    repasse_retido.aprovada = False
    repasse_retido.save()
    return HttpResponseRedirect('/tbl_repasse_retido.html')


def download_planilha_repasses_aprovados(request, *args, **kwargs):
    data_inicio = kwargs.get('data_inicio')
    data_fim = kwargs.get('data_fim')

    repasses_aprovados = RepasseAprovado.objects.filter(
        data_aprovado__range=[data_inicio, data_fim],
    )

    #exporte essa consulta para download
    # Crie um objeto Workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    

    # Adicione os cabeçalhos das colunas
    ws.append([
        'Vendedor ID',
        'Nome do Vendedor',
        'Data Consulta Inicial',
        'Data Consulta Final',
        'Total Repasse Retido',
        'Total Créditos',
        'Total Débitos',
        'Total Taxas',
        'Total Repasses'
    ])

    # Adicione os dados da consulta à planilha
    for repasse in repasses_aprovados:
        ws.append([
            repasse.cliente.id,
            repasse.cliente.nome,
            repasse.data_inicial,
            repasse.data_final,
            repasse.total_repasses_retidos(),
            repasse.total_creditos(),
            repasse.total_debitos(),
            repasse.total_taxas(),
            repasse.total_repasse()
        ])

    # Crie um objeto de resposta com os cabeçalhos apropriados para download de arquivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=repasses_aprovados-{datetime.now().date()}.xlsx'

    # Salve o arquivo Excel no objeto de resposta
    wb.save(response)

    return response