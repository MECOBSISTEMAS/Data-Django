# -*- encoding: utf-8 -*-
"""
Copyright (c) 2019 - present AppSeed.us
"""
from django.utils.text import slugify
from datetime import datetime, date, timedelta
from django import template
from django.contrib.auth.decorators import login_required
#importe o method_decorator: 
from django.utils.decorators import method_decorator
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.db.models import Case, Sum, When, F, Q, IntegerField, DecimalField, OuterRef, Subquery, Value, Max
from django.db.models.functions import Coalesce, Cast
from django.views.generic import View
from django.template import loader
from django.urls import reverse
from django.shortcuts import render
import random
import ast
import tempfile
import os
import json
import openpyxl
from decimal import Decimal
from openpyxl.utils import get_column_letter

from . import funcoes

from .existing_models import Contratos, ContratoParcelas, Pessoas, Eventos
#from .forms import CAD_ClienteForm, Calculo_RepasseForm
from .models import Calculo_Repasse, CadCliente, Debito, Credito, Taxa, RepasseRetido, Dado, RepasseAprovado, ParcelaTaxa

class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        if isinstance(obj, date):
            return obj.strftime('%Y-%m-%d')
        if isinstance(obj, datetime):
            return obj.strftime('%Y-%m-%d %H:%M:%S')
        return super().default(obj)


def pegar_pessoa(pessoa_id):
    try:
        pessoa = Pessoas.objects.get(id=pessoa_id)
    except Exception as e:
        return HttpResponse('Pessoa não encontrada, erro: ' + str(e))
    return pessoa

@login_required(login_url="/login/")
def index(request):
    context = {'segment': 'index'}

    html_template = loader.get_template('home/index.html')
    return HttpResponse(html_template.render(context, request))

@method_decorator(login_required(login_url="/login/"), name='dispatch')
class IndexView(View):
    TEMPLATE_NAME:str = 'home/index.html'
    def get(self, request):
        context = {'segment': 'index'}

        return HttpResponse(context=context, template_name=self.TEMPLATE_NAME,request=request)


@login_required(login_url="/login/")
def pages(request):
    context:dict = {}

    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:
        request.session['serialized_data'] = None
        #request.session['prestacao_diaria_data'] = None
        load_template = request.path.split('/')[-1]

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        context['segment'] = load_template
        

        if load_template == 'cad_clientes_table_bootstrap.html':
            if request.method == 'POST':
                if 'novo-cliente-cadastro' in request.POST:
                    cliente_id = request.POST.get('cliente_id')
                    
                    nome = request.POST.get('nome')
                    email = request.POST.get('email')
                    
                    sim = request.POST.get('sim')
                    nao = request.POST.get('nao')
                    operacional = request.POST.get('operacional')
                    tcc = request.POST.get('tcc')
                    honorarios = request.POST.get('honorarios')
                    """ animal = request.POST.get('animal')
                    evento = request.POST.get('evento') """
                    informar_repasse = request.POST.get('informar_repasse')
                    
                    if nome and email:
                        pessoa = Pessoas.objects.create(nome=nome, email=email)
                        cliente_id = pessoa.id
                    elif nome:
                        pessoa = Pessoas.objects.create(nome=nome)
                        cliente_id = pessoa.id
                    else:
                        try:
                            pessoa = Pessoas.objects.get(id=cliente_id)
                        except Exception as e:
                            return HttpResponse('Cliente não encontrado, erro: ' + str(e))
                        if CadCliente.objects.filter(cliente_id=cliente_id).exists():
                            return HttpResponse('Cliente já cadastrado')
                    CadCliente.objects.create(
                        vendedor = pessoa, sim=sim, nao=nao, 
                        operacional=operacional, tcc=tcc, 
                        honorarios=honorarios, informar_repasse=informar_repasse
                    )
                        
            context['cad_clientes'] = CadCliente.objects.all()
            
        elif load_template == 'tbl_prestacao_diaria.html':
            if request.method == 'POST':
                if 'exportar-xlsx' in request.POST:
                    if not request.session.get('prestacao_diaria_data'):
                        return HttpResponse('Nenhum dado encontrado para exportar')
                    return funcoes.exportar_planilha_prestacao_diaria(request)
                elif 'exportar-pdf' in request.POST:
                    return funcoes.gerar_arquivo_pdf(request, context)
                elif 'filtrar-prestacao-diaria' in request.POST:
                    bancos = request.POST.get('bancos')
                    data = request.POST.get('data')
                    context['data_consultada'] = data
                    context['repasses_semanais'] = CadCliente.objects.filter(
                        repasse_semanal=True,
                    ).annotate(
                        total_repasses=Coalesce(
                            Subquery(
                                Dado.objects.filter(
                                    dt_credito=data,
                                    id_vendedor=OuterRef('vendedor_id')
                                ).values('id_vendedor').annotate(
                                    soma_repasses=Sum('repasses')
                                ).values('soma_repasses')
                            ),
                            0,
                            output_field=DecimalField()
                        )
                    ).values('total_repasses', 'vendedor__nome', 'vendedor__id')
                    
                    context['repasses'] = (
                        Dado.objects
                        .filter(dt_credito=data, banco=str(bancos).upper())
                        .values('id_contrato', 'vendedor')
                        .annotate(
                            repasses=Sum('repasses')
                        )
                        .order_by('vendedor')
                    )
                        
                    
                    context['valores_pagos_honorarios'] = Dado.objects.filter(dt_credito=data, 
                        banco=str(bancos).upper()).aggregate(
                            valores_pagos=Sum('vl_pago'),
                            honorarios=Sum('me')
                        )
                        
                    context['comissionistas_do_mes'] = Dado.objects.filter(
                        dt_credito=data,comissao__isnull=False
                        ).values('comissao').annotate(comissoes=Sum('op'))
                    
                    context['repasses_geral'] = Dado.objects.filter(dt_credito=data,
                        banco=str(bancos).upper()).aggregate(
                            repasses=Sum('repasses')
                        )
                        
                    context['taxas'] = float(Dado.objects.filter(dt_credito=data, banco=str(bancos).upper()).aggregate(taxas=Sum('taxas'))['taxas'] or 0)
                    repasses_geral = float(context['repasses_geral']['repasses'] or 0)
                    repasses_semanais_vendedores_totais = (sum([float(querie['total_repasses']) for querie in context['repasses_semanais']]))
                    repasses_geral_descontado = repasses_geral - repasses_semanais_vendedores_totais
                    
                    context['repasses_geral_descontado'] = float("{:.2f}".format(repasses_geral_descontado))
                    
                    request.session['prestacao_diaria_data'] = {
                        'repasses_semanais': json.dumps(list(context['repasses_semanais']), cls=CustomJSONEncoder),
                        'comissionistas_do_mes': json.dumps(list(context['comissionistas_do_mes']), cls=CustomJSONEncoder),
                        'valores_pagos_honorarios': json.dumps(list(context['valores_pagos_honorarios'].values()), cls=CustomJSONEncoder),
                        'comissionistas_do_mes': json.dumps(list(context['comissionistas_do_mes'].values()), cls=CustomJSONEncoder),
                        'repasses_geral_descontado': json.dumps(context['repasses_geral_descontado'], cls=CustomJSONEncoder),
                        'taxas': json.dumps(context['taxas'], cls=CustomJSONEncoder),
                    }

            
        elif load_template == 'form_elements.html':
            if request.method == "POST":
                data_inicio = request.POST.get('data_inicio')  # 2022-08-01:str
                data_fim = request.POST.get('data_fim')  # 2022-08-21:str
                context['sql'] =  list(Dado.objects \
                    .filter(dt_credito__range=(data_inicio, data_fim)) \
                    .values('id_vendedor', 'id_contrato', 'vendedor', 'id_comprador', 'comprador',
                            'parcelas_contrato', 'vl_pago', 'dt_vencimento', 'dt_credito', 'banco', 
                            'contrato', 'evento', 'deposito', 'calculo', 'taxas', 'adi', 'me', 'op', 
                            'repasses', 'comissao') \
                    .annotate(total_repasse=Sum('repasses')) \
                    .values('id_vendedor', 'id_contrato', 'vendedor', 'id_comprador', 'comprador', 
                            'parcelas_contrato', 'vl_pago', 'dt_vencimento', 'dt_credito', 'banco', 
                            'contrato', 'evento', 'deposito', 'calculo', 'taxas', 'adi', 'me', 'op', 
                            'repasses', 'comissao', 'total_repasse', 'id'))
                request.session['serialized_data'] = json.dumps(context['sql'], cls=CustomJSONEncoder)
        
        elif load_template == 'tbl_bootstrap.html':
            if request.method == 'POST':
                if 'filtrar-tabela-repasse-cliente' in request.POST:
                    data_inicio = request.POST.get('data-inicio')
                    data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
                    data_fim = request.POST.get('data-fim')
                    data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
                    context['data_inicio'] = data_inicio
                    context['data_final'] = data_fim
                    dados_dias = {}
                    for i in range((datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 1):
                        dia = datetime.strptime(data_inicio, '%Y-%m-%d') + timedelta(days=i)
                        dados_dias[f'{dia.day}/{dia.month}/{dia.year}'] = Sum(
                            Case(
                                When(dt_credito__day=dia.day, then=F('repasses')),
                                default=0,
                                output_field=DecimalField(decimal_places=2, max_digits=14, validators=[]),
                            ),
                        )

                    context['dias_de_consulta'] = list(range(1, (datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 2))
                    #context['dias_de_consulta'] = [(data_inicio_dt + timedelta(days=x)).day for x in range((data_fim_dt - data_inicio_dt).days + 1)]
                    context['dados_dias'] = dados_dias
                    
                    context['dados'] = Dado.objects.filter(
                        dt_credito__gte=data_inicio, 
                        dt_credito__lte=data_fim,
                        repasse_aprovado=False,
                    ).values(
                        'id_vendedor','vendedor', id_dado=Max('id')
                    ).annotate(
                        total_repasses_retidos=Coalesce(
                            Subquery(
                                RepasseRetido.objects.filter(
                                    cliente_id=OuterRef('id_vendedor'),
                                    dt_rep_retido__gte=data_inicio,
                                    dt_rep_retido__lte=data_fim,
                                ).values('cliente_id')
                                .annotate(total=Sum('vlr_rep_retido'))
                                .values('total'),
                                output_field=DecimalField(max_digits=8, decimal_places=2)
                            ),
                            Value(0, output_field=DecimalField(max_digits=8, decimal_places=2))
                        ),
                        **dados_dias,
                        total_credito=Coalesce(
                            Subquery(
                                ParcelaTaxa.objects.filter(
                                    id_vendedor=OuterRef('id_vendedor'),
                                    dt_vencimento__gte=data_inicio,
                                    dt_vencimento__lte=data_fim,
                                    aprovada=True
                                ).values('id_vendedor')
                                .annotate(total=Sum('repasse'))
                                .values('total'),
                                output_field=DecimalField(max_digits=8, decimal_places=2)
                            ),
                            Value(0, output_field=DecimalField(max_digits=8, decimal_places=2))
                        ),
                        total_taxa=Coalesce(
                            Subquery(
                                Taxa.objects.filter(
                                    cliente_id=OuterRef('id_vendedor'),
                                    dt_taxa__gte=data_inicio,
                                    dt_taxa__lte=data_fim,
                                    aprovada=True
                                ).values('cliente_id')
                                .annotate(total=Sum('taxas'))
                                .values('total'),
                                output_field=DecimalField(max_digits=8, decimal_places=2)
                            ),
                            Value(0, output_field=DecimalField(max_digits=8, decimal_places=2))
                        ),
                        total_debito=Coalesce(
                            Subquery(
                                ParcelaTaxa.objects.filter(
                                    id_comprador=OuterRef('id_vendedor'),
                                    dt_vencimento__gte=data_inicio,
                                    dt_vencimento__lte=data_fim,
                                    aprovada=True,
                                ).values('id_comprador').annotate(total=Sum('desconto_total')).values('total'),
                                output_field=DecimalField(max_digits=8, decimal_places=2)
                            ),
                        Value(0, output_field=DecimalField(max_digits=8, decimal_places=2))
                        ),
                        total_repasse= -F('total_taxa') - F('total_debito') + F('total_credito') + F('total_repasses_retidos') + Sum(F('repasses'))
                    ).order_by('id_vendedor')
                    """ Subquery(
                                Debito.objects.filter(
                                    cliente_id=OuterRef('id_vendedor'),
                                    dt_debitado__gte=data_inicio,
                                    dt_debitado__lte=data_fim,
                                ).values('cliente_id')
                                .annotate(total=Sum('vl_debito'))
                                .values('total'),
                                output_field=DecimalField(max_digits=8, decimal_places=2)
                            ),
                            Value(0, output_field=DecimalField(max_digits=8, decimal_places=2)) """

                    request.session['serialized_data'] = json.dumps(list(context['dados']), cls=CustomJSONEncoder)

                    tbody = "<tr>"
                    for dado in context['dados']:
                        tbody += '<td><a name="aprovar-repasse" id="aprovar-repasse" class="btn btn-success btn-sm" href="/aprovar_repasse/{}/{}/{}/{}/{}/{}/{}/{}">Aprovar Repasse</a></td>'.format(
                            dado['id_vendedor'],data_inicio,data_fim, dado['total_repasses_retidos'],dado['total_credito'],dado['total_taxa'],dado['total_debito'],dado['total_repasse'])
                        #tbody += f'<td><a name="aprovar-repasse" id="aprovar-repasse" class="btn btn-success btn-sm" href="/aprovar_repasse/{}/{data_inicio}/{data_fim}">Aprovar Repasse</a></td>'
                        tbody += f"<td>{dado['id_vendedor']}</td>"
                        tbody += f"<td>{dado['vendedor']}</td>"
                        tbody += f"<td>{dado['total_repasses_retidos']}</td>"
                        for dia in dados_dias:
                            tbody += f"<td>{dado[dia]}</td>"
                        tbody += f"<td>{dado['total_credito']}</td>"
                        tbody += f"<td>{dado['total_taxa']}</td>"
                        tbody += f"<td>{dado['total_debito']}</td>"
                        tbody += f"<td>{dado['total_repasse']}</td>"
                        tbody += "</tr>"
                    context['tbody'] = tbody
                    #return render(request, 'home/tbl_bootstrap.html', context=context)
        
        elif load_template == 'tbl_credito.html':
            if request.method == 'POST':
                if 'novo-credito' in request.POST:
                    credor = request.POST.get('credor')
                    pagador = request.POST.get('pagador')
                    valor = request.POST.get('valor')
                    data_credito = request.POST.get('data-credito')
                    descricao = request.POST.get('descricao')
                    if credor == pagador:
                        return HttpResponse("Credor e Pagador não podem ser iguais")
                    if pagador:
                        try:
                            pessoa_pagadora = Pessoas.objects.get(id=pagador)
                        except Pessoas.DoesNotExist:
                            return HttpResponse("Pessoa não encontrada, tem certeza que ela esta no banco de dados ?")
                        except Pessoas.MultipleObjectsReturned:
                            return HttpResponse("Mais de uma pessoa encontrada")
                        Debito.objects.create(cliente=pessoa_pagadora, vl_debito=valor, dt_debitado=data_credito, descricao=descricao)
                    try:
                        pessoa = Pessoas.objects.get(id=credor)
                    except Pessoas.DoesNotExist:
                        return HttpResponse("Pessoa não encontrada, tem certeza que ela esta no banco de dados ?")
                    except Pessoas.MultipleObjectsReturned:
                        return HttpResponse("Mais de uma pessoa encontrada")
                    Credito.objects.create(cliente=Pessoas.objects.get(id=credor), vl_credito=valor, dt_creditado=data_credito, descricao=descricao)
                elif 'filtrar-credito' in request.POST:
                    data_inicio = request.POST.get('data-inicio')
                    data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
                    data_fim = request.POST.get('data-fim')
                    data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
                    credito_dias = {}
                    for i in range((data_fim_dt - data_inicio_dt).days + 1):
                        dia = datetime.strptime(data_inicio, '%Y-%m-%d') + timedelta(days=i)
                        credito_dias[f'dia_{dia.day}'] = Sum(
                            Case(
                                When(dt_creditado__day=dia.day, then=F('vl_credito')),
                                default=0,
                                output_field=DecimalField(),
                            ),
                        )

                    context['dias'] = list(range(1, (data_fim_dt - data_inicio_dt).days + 2))
                    #context['dias_de_consulta'] = [(data_inicio_dt + timedelta(days=x)).day for x in range((data_fim_dt - data_inicio_dt).days + 1)]
                    context['creditos_dias'] = credito_dias

                    context['creditos'] = Credito.objects.filter(
                        dt_creditado__gte=data_inicio, 
                        dt_creditado__lte=data_fim,
                    ).values(
                        'cliente_id', 'cliente__nome',
                    ).annotate(
                        **credito_dias,
                        total_credito=Sum('vl_credito')
                    ).order_by('cliente_id')
                    

                    tbody = ""
                    for credito in context['creditos']:
                        tbody += "<tr>"
                        tbody += f"<td>{credito['cliente_id']}</td>"
                        tbody += f"<td>{credito['cliente__nome']}</td>"
                        for dia in credito_dias:
                            credito_dia = credito[f'{dia}']
                            tbody += f"<td>{credito_dia}</td>"
                        tbody += f"<td>{credito['total_credito']}</td>"
                        tbody += "</tr>"
                    context['tbody'] = tbody
                        
        elif load_template == 'tbl_repasses_aprovados.html':
            if request.method == 'POST':
                pass
            context['repasses_aprovados'] = RepasseAprovado.objects.all()
                    
        elif load_template == 'tbl_debito.html':
            if request.method == 'POST':
                if 'novo-debito' in request.POST:
                    pagador = request.POST.get('pagador')
                    credor = request.POST.get('credor')
                    valor = request.POST.get('valor')
                    data_debito = request.POST.get('data-debito')
                    descricao = request.POST.get('descricao')
                    if credor == pagador:
                        return HttpResponse("Credor e Pagador não podem ser iguais")
                    if credor:
                        try:
                            pessoa_credora = Pessoas.objects.get(id=credor)
                        except Pessoas.DoesNotExist:
                            return HttpResponse("Pessoa não encontrada, tem certeza que ela esta no banco de dados ?")
                        except Pessoas.MultipleObjectsReturned:
                            return HttpResponse("Mais de uma pessoa encontrada")    
                        Credito.objects.create(cliente=pessoa_credora, vl_credito=valor, dt_creditado=data_debito, descricao=descricao)
                    try:
                        pessoa_pagadora = Pessoas.objects.get(id=pagador)
                    except Pessoas.DoesNotExist:
                        return HttpResponse("Pessoa não encontrada, tem certeza que ela esta no banco de dados ?")
                    except Pessoas.MultipleObjectsReturned:
                        return HttpResponse("Mais de uma pessoa encontrada")
                    Debito.objects.create(cliente=pessoa_pagadora, vl_debito=valor, dt_debitado=data_debito, descricao=descricao)
                elif 'filtrar-debito' in request.POST:
                    data_inicio = request.POST.get('data-inicio')
                    data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
                    data_fim = request.POST.get('data-fim')
                    data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
                    debito_dias = {}
                    for i in range((datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 1):
                        dia = datetime.strptime(data_inicio, '%Y-%m-%d') + timedelta(days=i)
                        debito_dias[f'dia_{dia.day}'] = Sum(
                            Case(
                                When(dt_debitado__day=dia.day, then=F('vl_debito')),
                                default=0,
                                output_field=DecimalField(),
                            ),
                        )

                    context['dias'] = list(range(1, (datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 2))
                    #context['dias_de_consulta'] = [(data_inicio_dt + timedelta(days=x)).day for x in range((data_fim_dt - data_inicio_dt).days + 1)]
                    context['debitos_dias'] = debito_dias
                    
                    context['debitos'] = Debito.objects.filter(
                        dt_debitado__gte=data_inicio, 
                        dt_debitado__lte=data_fim,
                    ).values(
                        'cliente_id', 'cliente__nome'
                    ).annotate(
                        **debito_dias,
                        total_debito=Sum('vl_debito')
                    ).order_by('cliente_id')
                    
                    tbody = ""
                    for debito in context['debitos']:
                        tbody += "<tr>"
                        tbody += f"<td>{debito['cliente_id']}</td>"
                        tbody += f"<td>{debito['cliente__nome']}</td>"
                        for dia in debito_dias:
                            debito_dia = debito[f'{dia}']
                            tbody += f"<td>{debito_dia}</td>"
                        tbody += f"<td>{debito['total_debito']}</td>"
                        tbody += "</tr>"
                    context['tbody'] = tbody
                    
        elif load_template == 'tbl_taxas.html':
            if request.method == 'POST':
                if 'nova-taxa' in request.POST:
                    cliente_id = request.POST.get('cliente')
                    valor = request.POST.get('valor')
                    data_taxa = request.POST.get('data-taxa')
                    descricao = request.POST.get('descricao')
                    tipo = request.POST.get('tipo')
                    Taxa.objects.create(cliente=Pessoas.objects.get(id=cliente_id), taxas=valor, dt_taxa=data_taxa, descricao=descricao, tipo=tipo)
                if 'filtrar-taxa' in request.POST:
                    data_inicio = request.POST.get('data-inicio')
                    data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
                    data_fim = request.POST.get('data-fim')
                    data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
                    context['data_inicio'] = data_inicio
                    context['data_fim'] = data_fim
                    taxas_dias = {}
                    for i in range((datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 1):
                        dia = datetime.strptime(data_inicio, '%Y-%m-%d') + timedelta(days=i)
                        taxas_dias[f'dia_{dia.day}'] = Sum(
                            Case(
                                When(dt_taxa__day=dia.day, then=F('taxas')),
                                default=0,
                                output_field=DecimalField(),
                            ),
                        )

                    context['dias'] = list(range(1, (datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 2))
                    #context['dias_de_consulta'] = [(data_inicio_dt + timedelta(days=x)).day for x in range((data_fim_dt - data_inicio_dt).days + 1)]
                    context['taxas_dias'] = taxas_dias
                    
                    context['taxas'] = Taxa.objects.filter(
                        dt_taxa__gte=data_inicio, 
                        dt_taxa__lte=data_fim,
                    ).values(
                        'cliente_id', 'cliente__nome'
                    ).annotate(
                        **taxas_dias,
                        total_taxa=Sum('taxas')
                    ).order_by('cliente_id')
                    
                    tbody = ""
                    for taxa in context['taxas']:
                        tbody += "<tr>"
                        tbody += f"<td>{taxa['cliente_id']}</td>"
                        tbody += f"<td>{taxa['cliente__nome']}</td>"
                        for dia in taxas_dias:
                            taxa_dia = taxa[f'{dia}']
                            tbody += f"<td>{taxa_dia}</td>"
                        tbody += f"<td>{taxa['total_taxa']}</td>"
                        tbody += "</tr>"
                    context['tbody'] = tbody
                    
                    context['taxas_nao_aprovadas'] = Taxa.objects.filter(
                        dt_taxa__gte=data_inicio,
                        dt_taxa__lte=data_fim,
                        aprovada=False,
                    )
                    
                    context['taxas_aprovadas'] = Taxa.objects.filter(
                        dt_taxa__gte=data_inicio,
                        dt_taxa__lte=data_fim,
                        aprovada=True,
                    )

                    
        elif load_template == 'tbl_repasse_retido.html':
            if request.method == 'POST':
                if 'novo-repasse-retido' in request.POST:
                    cliente_id = request.POST.get('cliente')
                    valor = request.POST.get('valor')
                    data_repasse_retido = request.POST.get('data-repasse-retido')
                    tipo = request.POST.get('tipo')
                    try:
                        cliente = Pessoas.objects.get(id=cliente_id)
                        RepasseRetido.objects.create(cliente=cliente, vlr_rep_retido=valor, dt_rep_retido=data_repasse_retido, tipo=tipo)
                    except Pessoas.DoesNotExist:
                        return HttpResponse("Cliente não encontrado")
                    except Pessoas.MultipleObjectsReturned:
                        return HttpResponse("Mais de um cliente encontrado, {}".format(Pessoas.objects.filter(id=cliente_id)))
                elif 'filtrar-repasse-retido' in request.POST:
                    data_inicio = request.POST.get('data-inicio')
                    data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
                    data_fim = request.POST.get('data-fim')
                    data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
                    repasses_retidos_dias = {}
                    for i in range((datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 1):
                        dia = datetime.strptime(data_inicio, '%Y-%m-%d') + timedelta(days=i)
                        repasses_retidos_dias[f'dia_{dia.day}'] = Sum(
                            Case(
                                When(dt_rep_retido__day=dia.day, then=F('vlr_rep_retido')),
                                default=0,
                                output_field=DecimalField(),
                            ),
                        )

                    context['dias'] = list(range(1, (datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 2))
                    #context['dias_de_consulta'] = [(data_inicio_dt + timedelta(days=x)).day for x in range((data_fim_dt - data_inicio_dt).days + 1)]
                    context['repasses_retidos_dias'] = repasses_retidos_dias
                    
                    context['repasses_retidos'] = RepasseRetido.objects.filter(
                        dt_rep_retido__gte=data_inicio, 
                        dt_rep_retido__lte=data_fim,
                    ).values(
                        'cliente_id', 'cliente__nome'
                    ).annotate(
                        **repasses_retidos_dias,
                        total_repasse_retido=Sum('vlr_rep_retido')
                    ).order_by('cliente_id')
                    
                    tbody = ""
                    for repasse_retido in context['repasses_retidos']:
                        tbody += "<tr>"
                        tbody += f"<td>{repasse_retido['cliente_id']}</td>"
                        tbody += f"<td>{repasse_retido['cliente__nome']}</td>"
                        for dia in repasses_retidos_dias:
                            repasse_retido_dia = repasse_retido[f'{dia}']
                            tbody += f"<td>{repasse_retido_dia}</td>"
                        tbody += f"<td>{repasse_retido['total_repasse_retido']}</td>"
                        tbody += "</tr>"
                    context['tbody'] = tbody

        elif load_template == 'tbl_resultado_financeiro.html':
            if request.method == 'POST':
                data_inicio = request.POST.get('data-inicio')
                data_fim = request.POST.get('data-fim')
                financeiro_dias = {}
                for i in range((datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 1):
                    dia = datetime.strptime(data_inicio, '%Y-%m-%d') + timedelta(days=i)
                    financeiro_dias[f'dia_{dia.day}'] = Sum(
                        Case(
                            When(dt_vencimento__day=dia.day, then=F('repasse')),
                            default=0,
                            output_field=DecimalField(decimal_places=2, max_digits=14, validators=[]),
                        ),
                    )
                context['dias'] = list(range(1, (datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 2))
                context['financeiro_dias'] = financeiro_dias
                context['valores_financeiros'] = ParcelaTaxa.objects.filter(
                    dt_vencimento__range=[data_inicio, data_fim]
                ).values('id_contrato', 'comprador', 'vendedor').annotate(
                    **financeiro_dias,
                )
                context['soma_valores_financeiro'] = ParcelaTaxa.objects.filter(
                    dt_vencimento__range=[data_inicio, data_fim]
                ).aggregate(
                    total_valor=Sum(Coalesce('valor', 0, output_field=DecimalField(decimal_places=2, max_digits=12))),
                    total_tcc=Sum(Coalesce('tcc', 0,output_field=DecimalField(decimal_places=2, max_digits=12))),
                    total_desconto=Sum(Coalesce('desconto_total', 0,output_field=DecimalField(decimal_places=2, max_digits=12))),
                    total_honorarios=Sum(Coalesce('honorarios', 0,output_field=DecimalField(decimal_places=2, max_digits=12))),
                    total_repasse=Sum(Coalesce('repasse', 0,output_field=DecimalField(decimal_places=2, max_digits=12))),
                )
                context['valores_financeiro_filtro'] = ParcelaTaxa.objects.filter(dt_vencimento__range=[data_inicio, data_fim])
                
                tbody = ""
                for valor_financeiro in context['valores_financeiros']:
                    tbody += "<tr>"
                    #!tbody += f"<td>{valor_financeiro['id_contrato']}</td>"
                    for dia in financeiro_dias:
                        valor_financeiro_dia = valor_financeiro[f'{dia}']
                        tbody += f"<td>{valor_financeiro_dia}</td>"
                    tbody += "</tr>"
                context['tbody_repasses'] = tbody
                
                context['tcc_filtro'] = ParcelaTaxa.objects.filter(
                    dt_vencimento__range=[data_inicio, data_fim], tcc__isnull=False
                    ).values('id_contrato', 'comprador', 'vendedor').annotate(
                        **funcoes.construir_dias_filtro(data_inicio, data_fim, 0, 'tcc'),
                    )
                tbody = ""
                for tcc in context['tcc_filtro']:
                    tbody += "<tr>"
                    for dia in financeiro_dias:
                        tcc_dia = tcc[f'{dia}']
                        tbody += f"<td>{tcc_dia}</td>"
                    tbody += "</tr>"
                context['tbody_tcc'] = tbody
                
                context['desconto_total_filtro'] = ParcelaTaxa.objects.filter(
                    dt_vencimento__range=[data_inicio, data_fim], desconto_total__isnull=False
                    ).values('id_contrato', 'comprador', 'vendedor').annotate(
                        **funcoes.construir_dias_filtro(data_inicio, data_fim, 0, 'desconto_total'),
                    )
                tbody = ""
                for desconto_total in context['desconto_total_filtro']:
                    tbody += "<tr>"
                    for dia in financeiro_dias:
                        desconto_total_dia = desconto_total[f'{dia}']
                        tbody += f"<td>{desconto_total_dia}</td>"
                    tbody += "</tr>"
                context['tbody_desconto_total'] = tbody
                
                context['honorarios_filtro'] = ParcelaTaxa.objects.filter(
                    dt_vencimento__range=[data_inicio, data_fim], honorarios__isnull=False
                    ).values('id_contrato', 'comprador', 'vendedor').annotate(
                        **funcoes.construir_dias_filtro(data_inicio, data_fim, 0, 'honorarios'),
                    )
                tbody = ""
                for honorarios in context['honorarios_filtro']:
                    tbody += "<tr>"
                    for dia in financeiro_dias:
                        honorarios_dia = honorarios[f'{dia}']
                        tbody += f"<td>{honorarios_dia}</td>"
                    tbody += "</tr>"
                context['tbody_honorarios'] = tbody
                    

        elif load_template == 'tbl_parcela_taxas.html':
            if request.method == 'POST':
                if 'filtrar-parcela-taxa' in request.POST:
                    data_inicio = request.POST.get('data-inicio')
                    data_fim = request.POST.get('data-fim')
                    context['data_inicio'] = data_inicio
                    context['data_fim'] = data_fim
                    context['parcelas_taxas'] = ParcelaTaxa.objects.filter(dt_vencimento__range=(data_inicio,data_fim), aprovada=False)
                    
        elif load_template == 'tbl_taxas_aprovadas.html':
            if request.method == 'POST':
                if 'filtrar-parcela-taxa-aprovada' in request.POST:
                    data_inicio = request.POST.get('data-inicio')
                    data_fim = request.POST.get('data-fim')
                    context['parcelas_taxas_aprovadas'] = ParcelaTaxa.objects.filter(dt_vencimento__range=(data_inicio, data_fim) , aprovada=True)
            context['parcelas_taxas_aprovadas'] = ParcelaTaxa.objects.filter(aprovada=True)
        
        elif load_template == 'pessoa_info.html':
            if request.method == 'POST':
                pass
            context['pessoa'] = Pessoas.objects.get(id=733 or None)
                
            

        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:
        pass

        """ html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request)) """

    """ except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request)) """





def upload_planilha_quinzenal(request, *args, **kwargs):
    if request.method == 'POST':
        planilha = request.FILES.get('docpicker')
        return HttpResponse(planilha)
    return HttpResponseRedirect('/tbl_credito_cessao.html')

def download_planilha_repasses(request, *args, **kwargs):
    if request.session.get('serialized_data') is None:
        return HttpResponse('Nenhum dado encontrado para download')
    with tempfile.TemporaryDirectory() as tmpdir:
        filepath = os.path.join(tmpdir, f'planilha_consulta_quinzenal_{slugify(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'ID Vendedor'
        sheet['B1'] = 'Nome Vendedor'
        sheet['C1'] = 'Valor Repasse Retido'
        
        # define a largura das primeiras colunas
        sheet.column_dimensions[get_column_letter(1)].width = 15
        sheet.column_dimensions[get_column_letter(2)].width = 30
        sheet.column_dimensions[get_column_letter(3)].width = 30
        
        for i, row in enumerate(json.loads(request.session.get('serialized_data'))):
            for j, value in enumerate(row):
                sheet.cell(row=i+2, column=j+1, value=row[value])
        workbook.save(filepath)
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(filepath)}"'
            return response

def download_planilha_cob(request, *args, **kwargs):
    if request.session.get('serialized_data') is None:
        return HttpResponse('Nenhum dado encontrado para download')
    with tempfile.TemporaryDirectory() as tmpdir:
        filepath = os.path.join(tmpdir, f'planilha_consulta_cob_{slugify(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        

        lista_header_consulta_sql = ['id_vendedor', 'id_contrato', 'vendedor', 'id_comprador', 'comprador',
                            'parcelas_contrato', 'vl_pago', 'dt_vencimento', 'dt_credito', 'banco',
                            'contrato', 'evento', 'deposito', 'calculo', 'taxas', 'adi', 'me', 'op',
                            'repasses', 'comissao', 'total_repasse', 'id']
        for i, value in enumerate(lista_header_consulta_sql):
            sheet[get_column_letter(i+1) + '1'] = str(value)
        
        for i, row in enumerate(json.loads(request.session.get('serialized_data'))):
            for j, header in enumerate(lista_header_consulta_sql):
                value = row.get(header, '')  # obtém o valor do dicionário usando a chave do cabeçalho
                sheet.cell(row=i+2, column=j+1, value=str(value))
        workbook.save(filepath)
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(filepath)}"'
            return response

def upload_planilha_cob(request, *args, **kwargs):
    if request.method == 'POST':
        planilha = request.FILES.get('docpicker')
        #verificar se o arquivo é do tipo xlsx
        if planilha.name.endswith('.xlsx'):
            #arquivo esta recebendo com sucesso, azer os devidos tratamentos
            wb = openpyxl.load_workbook(planilha)
            #nesse arquivo de planilha, a primeira aba é a que contem os dados, e so existe uma aba
            sheet = wb.worksheets[0]
            linha = 0
            for row in sheet.iter_rows(values_only=True):
                if linha < 1:
                    linha += 1
                    continue
                linha += 1
                if (row[0] and row[1] and row[2]) == None:
                    break
            
            
        #arquivo esta recebendo com sucesso, azer os devidos tratamentos
        return HttpResponse(planilha)
    return HttpResponseRedirect('/form_elements.html')

def upload_planilha_cavalos_cob(request, *args, **kwargs):
    if request.method == 'POST':
        planilha = request.FILES.get('docpicker')
        if planilha is None:
            return HttpResponse('Nenhum arquivo selecionado')
        elif planilha.name.endswith('.xlsx'):
            pessoa_nula = Pessoas.objects.get(id=99999)
            erros:list[str] = []
            linhas_nulas = 0
            wb = openpyxl.load_workbook(planilha)
            cob = wb.active
            linha = 0
            for row in cob.iter_rows(values_only=True):
                if linha < 1:
                    linha += 1
                    continue
                if (row[0] and row[1] and row[2]) == (None or ''):
                    linhas_nulas += 1
                    if linhas_nulas == 2:
                        break
                    continue
                codigo_vendedor = row[0]
                contrato_id = row[1]
                nome_vendedor = row[2]
                nome_comprador = row[3]
                parcela_paga = row[4]
                contrato_parcelas = row[5]
                valor_parcela = row[6]
                dt_vencimento = row[7]#dia/mes/ando
                """ dt_vencimento_dt = datetime.combine(date(day=dt_vencimento.day, 
                    month=dt_vencimento.month, 
                    year=dt_vencimento.year), datetime.min.time()) """
                dt_credito = row[8]
                """ dt_credito_dt = datetime.combine(date(day=dt_credito.day,
                    month=dt_credito.month,
                    year=dt_credito.year), datetime.min.time()) """
                banco = row[9]
                produto_ou_contrato = row[10]
                evento = row[11]
                deposito = row[12]
                calc = row[13]
                taxas = row[14]
                adi = row[15]
                me = row[16]
                op = row[17]
                repasses = row[18]
                comissao = row[19]
                try:
                    vendedor = Pessoas.objects.get(id=codigo_vendedor)
                except Pessoas.DoesNotExist:
                    vendedor = Pessoas.objects.create(id=codigo_vendedor, nome=nome_vendedor, email=f"{row[0]}_{row[2]}@gmail.com")
                except Pessoas.MultipleObjectsReturned:
                    erros.append(f"O vendedor {nome_vendedor} possui mais de um cadastro. linha: {linha}")
                    continue
                try:
                    comprador = Pessoas.objects.get(nome=nome_comprador)
                except Pessoas.DoesNotExist:
                    comprador = Pessoas.objects.create(nome=nome_comprador, email=f"{nome_comprador}@gmail.com")
                except Pessoas.MultipleObjectsReturned or Exception as e:
                    erros.append(f"O comprador {nome_comprador} possui mais de um cadastro. linha: {linha}")
                    continue
                    
                if row[1] is not None and type(row[1]) == int:
                    if row[1] == 1:
                        eventos = Eventos.objects.create(nome=row[4], leiloeiro=pessoa_nula, dt_evento=date.today(), tipo="qualquer")
                        Contratos.objects.create(vendedor=vendedor, comprador=comprador, eventos=eventos, descricao=row[10], pessoas_id_inclusao=pessoa_nula)
                    else:
                        try:
                            contratos = Contratos.objects.get(id=row[1])
                            eventos = Eventos.objects.get(id=contratos.eventos.id)
                            eventos.nome = evento
                            contratos.eventos = eventos
                            contratos.descricao = produto_ou_contrato
                            eventos.save()
                            contratos.save()
                        except Contratos.DoesNotExist:
                            eventos = Eventos.objects.create(nome=evento, leiloeiro=Pessoas.objects.get(id=99999),
                                dt_evento=date.today(), tipo="qualquer")
                            contratos = Contratos.objects.create(id=contrato_id,
                                vendedor=vendedor, comprador=comprador,
                                descricao=row[10], eventos=eventos, pessoas_id_inclusao=pessoa_nula)
                else:
                    erros.append(f"O contrato {row[1]} não existe para o comprador: {row[3]} e vendedor: {row[2]}. linha: {linha}")
                    continue
                try:
                    contrato_parcelas = ContratoParcelas.objects.get(contratos=contratos, 
                        nu_parcela=int(str(row[5].split('/')[0][1:]))
                        )
                    contrato_parcelas.dt_vencimento = row[7]
                    contrato_parcelas.vl_parcela = row[6]
                    contrato_parcelas.dt_credito = row[8]
                    contrato_parcelas.save()
                except ContratoParcelas.DoesNotExist:
                    contrato_parcelas = ContratoParcelas.objects.create(
                        contratos=contratos,
                        dt_vencimento=row[7], vl_parcela=row[6], dt_credito=row[8],
                        nu_parcela=int(str(row[5].split('/')[0][1:]))
                    )
                except ContratoParcelas.MultipleObjectsReturned:
                    erros.append(f"O contrato {contratos.id} possui mais de uma parcela com o numero {row[5]}. linha: {linha}")
                    #!chances muito poucas de acontecer, mas se acontecer, o sistema vai pegar o primeiro objeto
                    pass
                try:
                    calculo_repasse = Calculo_Repasse.objects.get(contrato_parcelas=contrato_parcelas)
                    calculo_repasse.deposito = row[12]
                    calculo_repasse.calculo = row[13]
                    calculo_repasse.taxas = row[14]
                    calculo_repasse.adi = row[15]
                    calculo_repasse.me = row[16]
                    calculo_repasse.op = row[17]
                    calculo_repasse.repasses = row[18]
                    calculo_repasse.comissao = row[19]
                    calculo_repasse.dt_credito = contrato_parcelas.dt_credito
                    calculo_repasse.save()
                except Calculo_Repasse.DoesNotExist:
                    calculo_repasse = Calculo_Repasse.objects.create(contrato_parcelas=contrato_parcelas,
                        deposito=row[12], calculo=row[13], taxas=row[14], adi=row[15],
                        me=row[16], op=row[17], repasses=row[18], comissao=row[19], dt_credito=contrato_parcelas.dt_credito
                    )
                    
                linha += 1
            return HttpResponse('Planilha de cavalos recebida com sucesso, linhas lidas: {}, erros: {}'.format(linha,erros))
        else:
            return HttpResponse('Arquivo não é do tipo xlsx')

def upload_planilha_cad_clientes(request, *args, **kwargs):
    if request.method == 'POST':
        planilha = request.FILES.get('docpicker')
        if planilha is None:
            return HttpResponse('Nenhum arquivo selecionado')
        elif planilha.name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(planilha)
            cad_cliente = wb.active
            linha = 0
            erros:list[str] = []
            for row in cad_cliente.iter_rows(values_only=True):
                if linha < 1:
                    linha += 1
                    continue
                #* sistema de parada
                if (row[0]) == None:
                    break
                nome_vendedor = row[0]
                id_vendedor = row[1]
                sim = row[2]
                nao = row[3]
                operacional = row[4]
                tcc = row[5]
                honorarios = row[6]
                repasse_semanal = row[7]
                try:
                    vendedor = Pessoas.objects.get(id=id_vendedor)
                except Pessoas.DoesNotExist:
                    vendedor = Pessoas.objects.create(id=id_vendedor, nome=nome_vendedor)
                except Pessoas.MultipleObjectsReturned:
                    erros.append('Erro ao criar o vendedor, Multiplas Pessoas Encontradas: {}, linha: {}'.format(nome_vendedor, linha))
                    continue
                try:
                    cad_cliente = CadCliente.objects.get(vendedor__id=vendedor.id)
                    cad_cliente.sim = sim
                    cad_cliente.nao = nao
                    cad_cliente.operacional = operacional
                    cad_cliente.tcc = tcc
                    cad_cliente.honorarios = honorarios
                    cad_cliente.repasse_semanal = True if str(repasse_semanal) == 'SIM' else False
                    cad_cliente.save()
                except CadCliente.DoesNotExist:
                    cad_cliente = CadCliente.objects \
                        .create(vendedor=vendedor, 
                            sim=sim, nao=nao, operacional=operacional, 
                            tcc=tcc, honorarios=honorarios, 
                            repasse_semanal=True if repasse_semanal == 'SIM' else False)
                except CadCliente.MultipleObjectsReturned:
                    erros.append('Erro ao criar o cadastro do cliente, Multiplos CadCliente,s Encontrados: {}, linha: {}'.format(nome_vendedor, linha))
                    continue
                
            return HttpResponse('Planilha {} recebida com sucesso'.format(planilha.name))
        else:
            return HttpResponse('Arquivo não é do tipo xlsx')

def upload_planilha_parcelas_taxas(request, *args, **kwargs):
    if request.method == "POST":
        planilha = request.FILES.get('docpicker')
        if planilha is None:
            return HttpResponse('Nenhum arquivo selecionado')
        elif not planilha.name.endswith('.xlsx'):
            return HttpResponse('Arquivo não é do tipo xlsx')
        linhas_nulas = 0
        linhas = 0
        erros:list[str] = []
        wb = openpyxl.load_workbook(planilha)
        segunda_quinzena = wb.active
        for row in segunda_quinzena.iter_rows(values_only=True):
            if linhas < 1:
                linhas += 1
                continue
            if row[0] == None or row[0] == '':
                break
            id_contrato = row[0]
            id_comprador = row[1]
            nome_comprador = row[2]
            id_vendedor = row[3]
            nome_vendedor = row[4]
            parcela = row[5]
            dt_vencimento = row[6]#datetime.strptime(row[6], "<%d/%m/%Y>").date()
            valor = row[7]
            tcc = row[8]
            ted = row[9]
            desconto_total = row[10]
            honorarios = row[11]
            repasse = row[12]
            ParcelaTaxa.objects.create(
                id_contrato=id_contrato, id_comprador=id_comprador, comprador=nome_comprador,
                id_vendedor=id_vendedor, vendedor=nome_vendedor, parcela=parcela, dt_vencimento=dt_vencimento,
                valor=valor, tcc=tcc, desconto_total=desconto_total, honorarios=honorarios, repasse=repasse
            )
            linhas += 1
        
        return HttpResponse("Planilha Recebida com sucesso, linhas lidas, {}, erros: {}".format(linhas, erros))

def upload_planilha_dados_brutos(request):
    if request.method == "POST":
        if request.FILES.get('docpicker') is None:
            return HttpResponse("Nenhum arquivo selecionado")
        if not request.FILES.get('docpicker').name.endswith('.xlsx'):
            return HttpResponse("Arquivo não é do tipo xlsx")
        planilha = request.FILES.get('docpicker')
        wb = openpyxl.load_workbook(planilha)
        cob = wb.active
        linhas_nulas = 0
        dados_modificados = 0
        dados_criados = 0
        linhas = 0
        erros:list[str] = []
        for row in cob.iter_rows(values_only=True):
            if linhas < 1:
                linhas += 1
                continue
            if (row[0] == None or row[0] == '') and (row[1] == None or row[1] == ''):
                linhas_nulas += 1
                if linhas_nulas > 1:
                    break
                continue
            linhas += 1
            vendedor_id = row[0]
            contrato_id = row[1]
            nome_vendedor = row[2]
            comprador = row[3]
            parcela_paga = row[4]
            parcelas_contrato = row[5]
            valor = row[6]
            data_vencimento = row[7]
            data_credito = row[8]
            banco = row[9]
            contrato = row[10]
            evento = row[11]
            deposito = row[12]
            calc = row[13]
            taxas = row[14]
            adi = row[15]
            me = row[16]
            op = row[17]
            repasses = row[18]
            comissao = row[19]
            id_excel = row[21]
            
            try:
                dado = Dado.objects.get(id_vendedor=vendedor_id, id_contrato=contrato_id,nu_parcela=parcela_paga, comprador=comprador, vl_pago=valor, contrato=contrato, evento=evento)
                """ altere todos os dados para a atual linha, exceto os id_vendedor=vendedor_id, id_contrato=contrato_id, comprador=comprador, vl_pago=valor """
                #dado.nu_parcela = parcela_paga
                dado.parcelas_contrato = parcelas_contrato
                dado.dt_vencimento = data_vencimento
                dado.dt_credito = data_credito
                dado.banco = banco
                #dado.contrato = contrato
                #dado.evento = evento
                dado.deposito = deposito
                dado.calculo = calc
                dado.taxas = taxas
                dado.adi = adi
                dado.me = me
                dado.op = op
                dado.repasses = repasses
                dado.comissao = comissao
                dado.save()
                dados_modificados += 1
            except Dado.DoesNotExist:
                Dado.objects.create(
                id_vendedor=vendedor_id,
                id_contrato=contrato_id,
                vendedor=nome_vendedor,
                comprador=comprador,
                nu_parcela=parcela_paga,
                parcelas_contrato=parcelas_contrato,#(1/2) :str, pegar somente o primeiro digte antes da barra
                vl_pago=valor,
                dt_vencimento=data_vencimento,
                dt_credito=data_credito,
                banco=banco,
                contrato=contrato,
                evento=evento,
                deposito=deposito,
                calculo=calc,
                taxas=taxas,
                adi=adi,
                me=me,
                op=op,
                repasses=repasses,
                comissao=comissao
            )
                dados_criados += 1
            except Dado.MultipleObjectsReturned:
                erros.append(f"Foi encontrado mais de um dado com os mesmos dados na linha {linhas}\n chaves primarias de busca são: id_vendedor={vendedor_id}, id_contrato={contrato_id}, comprador={comprador}, vl_pago={valor}")
            except Exception as e:
                erros.append(f"Erro na linha {linhas}, Exception Error:{e}")
            
            
        erros_html = "<ul>"
        for erro in erros:
            erros_html += f"<li>{erro}</li>"
        erros_html = "</ul>"
        return HttpResponse("Planilha recebida com sucesso <br> linhas lidas: {} <br> Dados Criados: {} <br> Dados Modificados : {} <br> erros: {}".format(linhas,dados_criados,dados_modificados, erros_html))
    return HttpResponse("HTTP REQUEST")

def aprovar_repasse(request, *args, **kwargs) :
    id_vendedor = kwargs.get('id_vendedor')
    data_inicial = kwargs.get('data_inicial')
    data_final = kwargs.get('data_final')
    total_repasses_retidos = kwargs.get('total_repasses_retidos')
    total_credito = kwargs.get('total_credito')
    total_debito = kwargs.get('total_debito')
    total_taxa = kwargs.get('total_taxa')
    total_repasse = kwargs.get('total_repasse')
    try:
        cliente = Pessoas.objects.get(id=id_vendedor)
    except Pessoas.DoesNotExist:
        cliente = Pessoas.objects.create(id=id_vendedor)
    except Pessoas.MultipleObjectsReturned:
        return HttpResponse(f"Erro ao criar cliente, mais de um cliente com o mesmo id, {id_vendedor}")
    dados = Dado.objects.filter(id_vendedor=id_vendedor, dt_credito__range=(data_inicial, data_final))

    repasse_aprovado = RepasseAprovado.objects.create(
        cliente=cliente,
        total_repasses_retidos=total_repasses_retidos,
        total_credito=total_credito,
        total_debito=total_debito,
        total_taxa=total_taxa,
        total_repasse=total_repasse,
        data_inicial=data_inicial,
        data_final=data_final,
    )
    for dado in dados:
        dado.repasse_aprovado = True
        dado.save()
        repasse_aprovado.dado.add(dado)
    #dados_nao_aprovados = Dado.objects.filter(dt_credito__range=(data_inicial, data_final), repasse_aprovado=False)
    dados = funcoes.repasses_nao_aprovados(data_inicial, data_final)
    
    
    return JsonResponse(
        {
        'dados':list(dados[0]),
        'dados_dias':list(dados[1]),
        'data_inicio':data_inicial,
        'data_fim':data_final, 
        'status':200
        }
    )
    dados_consultados_string = kwargs.get('dados_consultados')
    #dados_consultados_dict = ast.literal_eval(dados_consultados_string)
    data_inicial = kwargs.get('data_inicial')
    data_final = kwargs.get('data_final')
    id_vendedor = kwargs.get('id_vendedor')
    dados_consultados_string = dados_consultados_string.replace("Decimal('", "'").replace("')", "'")
    dados_consultados_dict = ast.literal_eval(dados_consultados_string)
    
    
    try:
        cliente = Pessoas.objects.get(id=id_vendedor)
    except Pessoas.DoesNotExist:
        cliente = Pessoas.objects.create(id=id_vendedor)
    except Pessoas.MultipleObjectsReturned:
        return HttpResponse(f"Erro ao criar cliente, mais de um cliente com o mesmo id, {dados_consultados_dict.get('id_vendedor')}")
    
    dados = Dado.objects.filter(id_vendedor=id_vendedor, dt_credito__range=(data_inicial, data_final))
    repasse_aprovado = RepasseAprovado.objects.create(
        cliente=cliente,
        total_repasses_retidos=dados_consultados_dict.get('total_repasses_retidos'),
        total_credito=dados_consultados_dict.get('total_credito'),
        total_debito=dados_consultados_dict.get('total_debito'),
        total_taxa=dados_consultados_dict.get('total_taxa'),
        total_repasse=dados_consultados_dict.get('total_repasse'),
        data_inicial=data_inicial,
        data_final=data_final,
    )
    for dado in dados:
        dado.repasse_aprovado = True
        dado.save()
        repasse_aprovado.dado.add(dado)
    #return HttpResponseRedirect('/tbl_bootstrap.html')
    dados_nao_aprovados = Dado.objects.filter(dt_credito__range=(data_inicial, data_final), repasse_aprovado=False).values()
    dados_nao_aprovados_json = json.dumps(list(dados_nao_aprovados))
    return JsonResponse({'dados': dados_nao_aprovados_json, 'data_inicial': data_inicial, 'data_final': data_final, 'status': 200})

def desaprovar_repasse(request, *args, **kwargs):
    repasse_aprovado = RepasseAprovado.objects.get(id=kwargs.get('repasse_aprovado_id'))
    for dado in repasse_aprovado.dado.all():
        dado.repasse_aprovado = False
        dado.save()
    repasse_aprovado.delete()
    return HttpResponseRedirect('/tbl_repasses_aprovados.html')

def aprovar_parcela_taxa(request, *args, **kwargs):
    parcela_taxa = ParcelaTaxa.objects.get(id=kwargs.get('parcela_taxa_id'))
    data_inicio = kwargs.get('data_inicio')
    data_fim = kwargs.get('data_fim')
    parcela_taxa.aprovada = True
    parcela_taxa.save()
    
    return JsonResponse(
        {
            'parcelas_taxas': list(ParcelaTaxa.objects.filter(dt_vencimento__range=(data_inicio, data_fim), aprovada=False).values()),
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'status': 200,
        }
    )
    return HttpResponseRedirect('/tbl_parcela_taxas.html')

def desaprovar_parcela_taxa(request, *args, **kwargs):
    parcela_taxa_aprovada = ParcelaTaxa.objects.get(id=kwargs.get('parcela_taxa_id'))
    parcela_taxa_aprovada.aprovada = False
    parcela_taxa_aprovada.save()
    return HttpResponseRedirect('tbl_taxas_aprovadas.html')

def aprovar_taxa_manual(request, *args, **kwargs):
    taxa = Taxa.objects.get(id=kwargs.get('taxa_id'))
    taxa.aprovada = True
    taxa.save()
    return HttpResponseRedirect('/tbl_taxas.html')

def desaprovar_taxa_manual(request, *args, **kwargs):
    taxa = Taxa.objects.get(id=kwargs.get('taxa_id'))
    taxa.aprovada = False
    taxa.save()
    return HttpResponseRedirect('/tbl_taxas.html')