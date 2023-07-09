from datetime import datetime
from django.utils.text import slugify
import tempfile
from datetime import datetime, timedelta
from django.db.models import Sum, Case, When, F, DecimalField, OuterRef, Subquery, Value, Max
from django.db.models.functions import Coalesce
from reportlab.pdfgen import canvas
from django.http import HttpResponse
import json
import openpyxl
import os

from .models import Debito, Credito, Taxa, Dado, RepasseRetido


def construir_dias_filtro(data_inicio: str, data_fim: str, dt_vencimento, campo: str) -> dict:
    dias_context = {}
    for i in range((datetime.strptime(data_fim, '%Y-%m-%d') - datetime.strptime(data_inicio, '%Y-%m-%d')).days + 1):
        dia = datetime.strptime(data_inicio, '%Y-%m-%d') + timedelta(days=i)
        dias_context[f'dia_{dia.day}'] = Sum(
            Case(
                When(dt_vencimento__day=dia.day, then=F(campo)),
                default=0,
                output_field=DecimalField(
                    decimal_places=2, max_digits=14, validators=[]),
            ),
        )
    return dias_context


def gerar_arquivo_pdf(request, context):
    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        pdf = canvas.Canvas(tmp.name, initialFontSize=12,
                            initialFontName='Helvetica', pagesize=(595, 842))
        pdf.drawString(100, 750, "Relatório PDF do Prestação ia")
        #pdf.drawString(100, 700, "Adicione aqui os dados necessários")
        pdf.drawString(100, 650, request.session.get('comissionistas_do_mes'))
        pdf.drawString(100, 600, request.session.get('repasses_geral'))
        pdf.drawString(100, 550, request.session.get('repasses'))
        pdf.drawString(100, 500, request.session.get('comissoes'))
        pdf.drawString(100, 450, request.session.get(
            'valores_pagos_honorarios'))
        pdf.drawString(100, 400, request.session.get(
            'repasses_geral_descontado'))
        pdf.save()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="prestacao_ia_pdf_{slugify(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}.pdf"'
        response.write(tmp.read())
        return response


def exportar_planilha_prestacao_ia(request, *args, **kwargs):
    with tempfile.TemporaryDirectory() as tmpdir:
        filepath = os.path.join(tmpdir, f'planilha_prestacao_ia_{slugify(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        context_keys = request.session.get('prestacao_diaria_data').keys()
        context_values = [ json.loads(request.session.get('prestacao_diaria_data')[key]) for key in context_keys]
        ultima_linha = 0
        ultima_coluna = 0
        
        """ 'repasses_semanais
            'comissionistas_do_mes
            'valores_pagos_honorarios
            'comissionistas_do_mes
            'repasses_geral_descontado """
        
        valores_pagos_e_honorarios = json.loads(request.session.get('prestacao_diaria_data')['valores_pagos_honorarios'])
        sheet.cell(row=1, column=1, value='Valor Pago')
        sheet.cell(row=1, column=2, value='honorarios')
        sheet.cell(row=2, column=1, value=valores_pagos_e_honorarios[0])
        sheet.cell(row=2, column=2, value=valores_pagos_e_honorarios[1])
        ultima_coluna = 2
        ultima_linha = 2
        ultima_coluna += 1
        ultima_linha += 2
        sheet.cell(row=ultima_linha, column=1, value='Comissionistas')
        sheet.cell(row=ultima_linha, column=2, value='Comissoes')
        
        comissionistas_do_mes = json.loads(request.session.get('prestacao_diaria_data')['comissionistas_do_mes'])
        for comissionista_unico in comissionistas_do_mes:
            sheet.cell(row=ultima_linha+2, column=1, value=comissionista_unico['comissao'])
            sheet.cell(row=ultima_linha+2, column=2, value=comissionista_unico['comissoes'])
            ultima_linha += 1
        
        ultima_linha += 1
        sheet.cell(row=ultima_linha+1, column=1, value='Repasses Semanais')
        repasses_semanais = json.loads(request.session.get('prestacao_diaria_data')['repasses_semanais'])
        for repasse_semanal in repasses_semanais:
            sheet.cell(row=ultima_linha+2, column=1, value=repasse_semanal['vendedor__id'])
            sheet.cell(row=ultima_linha+2, column=2, value=repasse_semanal['vendedor__nome'])
            sheet.cell(row=ultima_linha+2, column=3, value=repasse_semanal['total_repasses'])
            ultima_linha += 1
        ultima_linha += 2
        sheet.cell(row=ultima_linha, column=1, value='Repasses Geral')
        ultima_linha += 1
        sheet.cell(row=ultima_linha, column=1, value=json.loads(request.session.get('prestacao_diaria_data')['repasses_geral_descontado']))
        #context['taxas'] =float(Dado.objects.filter(dt_credito=data, banco=str(bancos).upper()).aggregate(taxas=Sum('taxas'))['taxas'] or 0)
        taxas = json.loads(request.session.get('prestacao_diaria_data')['taxas'])
        sheet.cell(row=ultima_linha, column=2, value=f"Taxas (não aplicada) {taxas}")
        
        
        
        workbook.save(filepath)
        with open(filepath, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(filepath)}"'
            return response
        
def repasses_nao_aprovados(data_inicial, data_final):
    """ Essa função ira retornar todos os dados não aprovadas porem  """
    dados_dias = {}
    for i in range((datetime.strptime(data_final, '%Y-%m-%d') - datetime.strptime(data_inicial, '%Y-%m-%d')).days + 1):
        dia = datetime.strptime(data_inicial, '%Y-%m-%d') + timedelta(days=i)
        dados_dias[f'{dia.day}/{dia.month}/{dia.year}'] = Sum(
            Case(
                When(dt_credito__day=dia.day, then=F('repasses')),
                default=0,
                output_field=DecimalField(decimal_places=2, max_digits=14, validators=[]),
            ),
        )

    dados = Dado.objects.filter(
        dt_credito__gte=data_inicial, 
        dt_credito__lte=data_final,
        repasse_aprovado=False,
    ).values(
        'id_vendedor','vendedor', id_dado=Max('id')
    ).annotate(
        total_repasses_retidos=Coalesce(
            Subquery(
                RepasseRetido.objects.filter(
                    cliente_id=OuterRef('id_vendedor'),
                    dt_rep_retido__gte=data_inicial,
                    dt_rep_retido__lte=data_final,
                ).values('cliente_id')
                .annotate(total=Sum('vlr_rep_retido'))
                .values('total'),
                output_field=DecimalField(max_digits=8, decimal_places=2)
            ),
            Value(0, output_field=DecimalField(max_digits=8, decimal_places=2))
        ),
        **dados_dias,
        total_credito=Coalesce(
            Subquery(
                Credito.objects.filter(
                    cliente_id=OuterRef('id_vendedor'),
                    dt_creditado__gte=data_inicial,
                    dt_creditado__lte=data_final,
                ).values('cliente_id')
                .annotate(total=Sum('vl_credito'))
                .values('total'),
                output_field=DecimalField(max_digits=8, decimal_places=2)
            ),
            Value(0, output_field=DecimalField(max_digits=8, decimal_places=2))
        ),
        total_taxa=Coalesce(
            Subquery(
                Taxa.objects.filter(
                    cliente_id=OuterRef('id_vendedor'),
                    dt_taxa__gte=data_inicial,
                    dt_taxa__lte=data_final,
                ).values('cliente_id')
                .annotate(total=Sum('taxas'))
                .values('total'),
                output_field=DecimalField(max_digits=8, decimal_places=2)
            ),
            Value(0, output_field=DecimalField(max_digits=8, decimal_places=2))
        ),
        total_debito=Coalesce(
            Subquery(
                Debito.objects.filter(
                    cliente_id=OuterRef('id_vendedor'),
                    dt_debitado__gte=data_inicial,
                    dt_debitado__lte=data_final,
                ).values('cliente_id')
                .annotate(total=Sum('vl_debito'))
                .values('total'),
                output_field=DecimalField(max_digits=8, decimal_places=2)
            ),
            Value(0, output_field=DecimalField(max_digits=8, decimal_places=2))
        ),
        total_repasse= -F('total_taxa') - F('total_debito') + F('total_credito') + F('total_repasses_retidos') + Sum(F('repasses'))
    )
    return [dados, dados_dias]